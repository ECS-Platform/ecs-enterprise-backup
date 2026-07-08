"""Predefined Queries — load controls from ECS_Query_Driven_Control_Library_Consolidated.xlsx."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from modules.operations.engines.predefined_query_audit import (
    get_execution_audit_log,
    get_execution_history_for_control,
    record_execution_audit,
)
from modules.operations.engines.predefined_query_evidence import (
    get_latest_evidence_for_control,
    prepare_evidence_record,
    register_with_evidence_repository,
)
from modules.shared.utils.pagination import paginate

EXCEL_FILENAME = "ECS_Query_Driven_Control_Library_Consolidated.xlsx"

HEADER_ALIASES: dict[str, list[str]] = {
    "control_id": ["control id", "controlid", "control_id", "id", "control ref", "control reference"],
    "control_name": ["control name", "controlname", "control_name", "name", "control title", "title"],
    "framework_coverage": [
        "framework coverage",
        "frameworks",
        "framework",
        "framework mapping",
        "framework(s)",
        "applicable frameworks",
    ],
    "query": ["query", "predefined query", "sql query", "command", "script", "check query", "validation query"],
    "description": ["description", "control description", "desc", "details", "requirement description"],
    "evidence_type": ["evidence type", "evidencetype", "evidence_type", "expected evidence", "evidence"],
}

TECHNOLOGY_RULES: list[tuple[str, list[str]]] = [
    ("GitLeaks", ["gitleaks detect", "gitleaks"]),
    ("Trivy", ["trivy image", "trivy"]),
    ("SonarQube", ["/api/issues/search", "sonarqube"]),
    # Aerospike — asinfo/asadm CLI surface (must precede generic matches).
    ("Aerospike", ["asinfo", "asadm", "aerospike"]),
    ("NGINX", ["nginx -t", "nginx -T", "nginx -v", "/etc/nginx", "sites-enabled", "ssl_protocols",
               "ssl_ciphers", "server_tokens"]),
    # YugabyteDB (YSQL) — must precede PostgreSQL because YSQL is PG-wire
    # compatible; only the yb_servers() signature is Yugabyte-specific in text.
    ("YugabyteDB", ["yb_servers(", "yb_servers", "yugabyte"]),
    # Aurora MySQL / MySQL — MySQL-specific SQL surface (SHOW VARIABLES, mysql.user,
    # SHOW DATABASES/PROCESSLIST). Placed before PostgreSQL to avoid "show " overlap.
    ("Aurora MySQL", ["show variables", "from mysql.user", "mysql.user", "show processlist",
                      "show databases", "have_ssl", "require_secure_transport"]),
    ("PostgreSQL", ["pg_stat_replication", "show ssl", "show password_encryption", "from pg_", " pg_",
                    "pg_database", "pg_roles", "pg_extension", "pg_tables"]),
    ("Oracle", ["dba_role_privs", "v$encryption_wallet", " v$", " dba_", "from dba_"]),
    ("SQL Server", ["@@version", "sys.server_principals", "sys.databases", "serverproperty(",
                    "sys.dm_database_encryption_keys", "sys.server_audit"]),
    ("Redis", ["redis-cli"]),
    ("Apache HTTPD", ["apachectl", "httpd -", "apache2 -", "/etc/httpd", "/etc/apache2"]),
    ("Tomcat", ["catalina", "tomcat-users.xml", "server.xml", "$catalina_home"]),
    ("OpenShift", ["oc get", "oc version", "clusteroperators", "get scc"]),
    ("Kubernetes", ["kubectl", "clusterrolebindings", "networkpolicies"]),
    ("Windows", ["get-hotfix", "get-mpcomputerstatus", "get-aduser", "powershell"]),
    ("Linux", ["df -h", "free -m", "timedatectl", "cat /etc/ssh/sshd_config", "/etc/ssh", "/etc/passwd",
               "/etc/group", "systemctl", "yum ", "apt-get", "dpkg", "rpm -", "hostname", "uptime"]),
]

ALLOWED_POSTGRESQL_QUERIES: frozenset[str] = frozenset({
    "show ssl;",
    "show password_encryption;",
    "select * from pg_stat_replication;",
    "select rolname, rolsuper, rolcreaterole, rolcreatedb, rolcanlogin from pg_roles;",
    "select datname, pg_database_size(datname) as size_bytes from pg_database;",
    "select datname, usename, application_name, client_addr, state from pg_stat_activity;",
    "select extname, extversion from pg_extension;",
    "select extname from pg_extension where extname in ('pgaudit');",
    "select name, setting from pg_settings where name in ('max_connections', "
    "'superuser_reserved_connections');",
    "select pid, usename, state, now() - query_start as runtime from pg_stat_activity "
    "where state <> 'idle' and now() - query_start > interval '5 minutes';",
    "select pg_postmaster_start_time() as start_time, now() - pg_postmaster_start_time() "
    "as uptime;",
    "select schemaname, count(*) as table_count from pg_tables where schemaname not in "
    "('pg_catalog', 'information_schema') group by schemaname;",
    "select name, setting from pg_settings where name in ('log_connections', "
    "'log_disconnections', 'ssl', 'password_encryption', 'log_statement');",
})

ALLOWED_YUGABYTE_QUERIES: frozenset[str] = frozenset({
    "select * from yb_servers();",
    "select version();",
    "select datname, usename, application_name, client_addr, state from pg_stat_activity;",
    "select rolname, rolsuper, rolcreaterole, rolcreatedb, rolcanlogin from pg_roles;",
    "select datname, pg_database_size(datname) as size_bytes from pg_database;",
    "select schemaname, tablename, tableowner from pg_tables where schemaname not in "
    "('pg_catalog', 'information_schema');",
    "select extname, extversion from pg_extension;",
    "show ssl;",
    "select name, setting from pg_settings where name in ('max_connections', "
    "'superuser_reserved_connections');",
    "select pid, usename, state, now() - query_start as runtime from pg_stat_activity "
    "where state <> 'idle' and now() - query_start > interval '5 minutes';",
    "select name, setting from pg_settings where name in ('log_connections', "
    "'log_disconnections', 'ssl', 'password_encryption');",
})

ALLOWED_MYSQL_QUERIES: frozenset[str] = frozenset({
    "show variables like 'have_ssl';",
    "show variables like 'require_secure_transport';",
    "show variables like 'log_bin';",
    "show variables like 'server_audit%';",
    "select user, host, plugin from mysql.user;",
    "select version();",
    "show databases;",
    "show processlist;",
    "select user, host, select_priv, insert_priv, update_priv, delete_priv, create_priv, "
    "drop_priv, super_priv from mysql.user;",
    "show variables like '%ssl%';",
    "show variables like 'max_connections';",
    "select id, user, host, db, time, state from information_schema.processlist "
    "where command <> 'sleep' and time > 300;",
    "show global status like 'aborted_connects';",
    "show global status like 'uptime';",
})

ALLOWED_ORACLE_QUERIES: frozenset[str] = frozenset({
    "select * from v$version;",
    "select name, open_mode, database_role from v$database;",
    "select wallet_type, status, wallet_order from v$encryption_wallet;",
    "select name, value from v$parameter where name = 'audit_trail';",
    "select profile, resource_name, limit from dba_profiles where resource_name in "
    "('failed_login_attempts','password_lock_time','password_life_time');",
    "select username, account_status, common, oracle_maintained from dba_users where "
    "username in ('sys','system') or account_status <> 'open';",
    "select grantee, granted_role, admin_option, default_role from dba_role_privs;",
    "select tablespace_name, status, contents from dba_tablespaces;",
    "select tablespace_name, encrypted from dba_tablespaces;",
    "select username, status, machine, program from v$session where username is not null;",
    "select resource_name, current_utilization, max_utilization, limit_value from "
    "v$resource_limit where resource_name in ('sessions', 'processes');",
    "select instance_name, status, startup_time from v$instance;",
    "select sid, username, status, last_call_et from v$session where username is not null "
    "and status = 'active' and last_call_et > 300;",
    "select owner, object_type, count(*) as object_count from dba_objects where owner not "
    "in ('sys','system') group by owner, object_type;",
})


def _norm_sql(q: str) -> str:
    return re.sub(r"\s+", " ", q.strip().lower()).rstrip(";") + ";"


# SQL Server + MongoDB allow-lists are derived from the supplementary catalog so
# they stay in sync automatically. (Built lazily below after the catalog import.)
try:
    from modules.operations.engines.supplementary_query_catalog import (
        SHELL_CONTROL_IDS as _SHELL_CONTROL_IDS,
        SQLSERVER_QUERIES as _SQLSERVER_QUERIES,
        MONGODB_QUERIES as _MONGODB_QUERIES,
        SUPPLEMENTARY_QUERY_BY_ID as _SUPPLEMENTARY_QUERY_BY_ID,
    )
except Exception:  # noqa: BLE001 - supplementary catalog is optional/additive
    _SUPPLEMENTARY_QUERY_BY_ID = {}
    _SHELL_CONTROL_IDS = frozenset()
    _SQLSERVER_QUERIES = []
    _MONGODB_QUERIES = []

ALLOWED_SQLSERVER_QUERIES: frozenset[str] = frozenset(
    _norm_sql(e["query"]) for e in _SQLSERVER_QUERIES
)
#: MongoDB "command specs" (catalog query text) enabled for live execution.
ALLOWED_MONGODB_COMMANDS: frozenset[str] = frozenset(
    (e["query"] or "").strip() for e in _MONGODB_QUERIES
)

# Curated set of controls wired for live execution. Extended with every
# supplementary DB query (PGX-*, YBX-*, MYX-*) so the new connectors are runnable.
LIVE_CONTROL_IDS: frozenset[str] = frozenset({
    "DB-001",
    "DB-002",
    "DB-003",
    "OS-001",
    "OS-002",
    "APP-001",
    "APP-002",
    "APPSEC-001",
    "APPSEC-002",
} | set(_SUPPLEMENTARY_QUERY_BY_ID.keys()))

SONAR_CONTROL_MODES: dict[str, str] = {
    "APP-001": "projects",
    "APP-002": "issues",
}

_controls: list[dict[str, Any]] = []
_validation_report: dict[str, Any] = {}
_loaded = False
_errors_found: list[str] = []
_errors_fixed: list[str] = []


def _connector_config_loaded() -> bool:
    from modules.operations.engines.query_connectors import CONNECTOR_CONFIG

    return bool(CONNECTOR_CONFIG)


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[3]


def _excel_path() -> Path:
    return _repo_root() / EXCEL_FILENAME


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _resolve_header_index(headers: list[str]) -> dict[str, int | None]:
    mapping: dict[str, int | None] = {key: None for key in HEADER_ALIASES}
    for idx, header in enumerate(headers):
        h = _normalize_header(header)
        if not h:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if mapping[field] is not None:
                continue
            if h in aliases or any(alias in h for alias in aliases):
                mapping[field] = idx
    return mapping


def _cell_value(row: tuple, index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def _parse_frameworks(raw: str) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[,;|/\n]+", raw)
    frameworks = []
    seen: set[str] = set()
    for part in parts:
        fw = part.strip()
        if fw and fw.lower() not in seen:
            seen.add(fw.lower())
            frameworks.append(fw)
    return frameworks


def detect_technology(query: str) -> str:
    """Deterministic technology detection from query text — no AI."""
    if not query or not query.strip():
        return ""
    q = query.lower()
    for technology, patterns in TECHNOLOGY_RULES:
        for pattern in patterns:
            if pattern.lower() in q:
                return technology
    return "Unknown"


# Explicit RHEL technology labels (shell checks via the Linux connector).
RHEL8_TECH = "Red Hat Enterprise Linux 8.x"
RHEL9_TECH = "Red Hat Enterprise Linux 9.x"
# Technologies executed as shell/CLI commands through a command connector
# (docker-exec Linux connector, or a local kubectl/oc subprocess).
_SHELL_TECHNOLOGIES: frozenset[str] = frozenset({
    "Linux", "NGINX", RHEL8_TECH, RHEL9_TECH,
    "Redis", "Apache HTTPD", "Tomcat", "Kubernetes", "OpenShift", "Aerospike",
})

# Technologies that have a real, executable connector implementation.
_IMPLEMENTED_CONNECTOR_TECH: frozenset[str] = frozenset(
    {"PostgreSQL", "YugabyteDB", "Aurora MySQL", "Oracle", "SQL Server", "MongoDB",
     "Linux", "NGINX", RHEL8_TECH, RHEL9_TECH,
     "Redis", "Apache HTTPD", "Tomcat", "Kubernetes", "OpenShift", "Aerospike",
     "SonarQube", "Trivy", "GitLeaks"}
)
# Technologies whose connector classes exist but are not yet runnable
# (SSHConnector raises NotImplementedError).
_GENERIC_CONNECTOR_TECH: frozenset[str] = frozenset({"Windows"})


def _dependency_available(technology: str) -> bool:
    """Return True when the runtime dependency for a technology is importable.

    Python-driver technologies gate on their import; subprocess/CLI technologies
    (shell containers, kubectl/oc) surface availability at execution time and are
    treated as dependency-available at capability-assessment time.
    """
    import importlib.util

    if technology in ("PostgreSQL", "YugabyteDB"):
        return importlib.util.find_spec("psycopg2") is not None
    if technology == "Aurora MySQL":
        return importlib.util.find_spec("pymysql") is not None
    if technology == "Oracle":
        return importlib.util.find_spec("oracledb") is not None
    if technology == "SQL Server":
        return importlib.util.find_spec("pyodbc") is not None
    if technology == "MongoDB":
        return importlib.util.find_spec("pymongo") is not None
    # Shell/CLI technologies (Linux/NGINX/RHEL/Redis/Apache/Tomcat/K8s/OpenShift)
    # run via docker exec or a local CLI; no importable dependency to gate here.
    return True


def assess_execution_capability(control: dict[str, Any]) -> dict[str, Any]:
    """Single source of truth for a control's *truthful* execution status.

    Returns ``{"executable": bool, "status": str, "reason": str}`` where status
    is one of: Manual, Unsupported Technology, Connector Missing, Configuration
    Required, Dependency Missing, Ready. "Ready" is returned ONLY when the
    control is genuinely executable (implemented connector + dependency present +
    wired for live execution) — never as a cosmetic label.
    """
    if not control.get("predefined"):
        return {"executable": False, "status": "Manual",
                "reason": "No predefined query — manual evidence collection required."}

    technology = control.get("technology") or ""
    if not technology or technology == "Unknown":
        return {"executable": False, "status": "Unsupported Technology",
                "reason": "Query text did not match any known technology pattern, so no connector can run it."}

    if technology in _GENERIC_CONNECTOR_TECH or technology not in _IMPLEMENTED_CONNECTOR_TECH:
        return {"executable": False, "status": "Connector Missing",
                "reason": f"No executable connector is implemented for {technology} yet."}

    if control.get("control_id") not in LIVE_CONTROL_IDS:
        return {"executable": False, "status": "Configuration Required",
                "reason": f"The {technology} connector exists but this control is not yet wired for live "
                          f"execution (target / query allow-list not configured)."}

    if not _dependency_available(technology):
        _driver_map = {
            "PostgreSQL": "psycopg2",
            "YugabyteDB": "psycopg2",
            "Aurora MySQL": "PyMySQL",
            "Oracle": "python-oracledb",
            "SQL Server": "pyodbc (+ an ODBC driver)",
            "MongoDB": "pymongo",
        }
        dep = _driver_map.get(technology, "the required driver")
        return {"executable": False, "status": "Dependency Missing",
                "reason": f"The {technology} driver ({dep}) is not installed in this environment."}

    return {"executable": True, "status": "Ready",
            "reason": f"{technology} connector is available and this control is enabled for live execution."}


def _derive_status(control: dict[str, Any]) -> str:
    return assess_execution_capability(control)["status"]


def _last_execution(control_id: str) -> str:
    history = get_execution_history_for_control(control_id, limit=1)
    if history:
        return history[0].get("execution_time", "—")
    return "Not Executed"


def _load_from_excel() -> tuple[list[dict[str, Any]], list[str]]:
    path = _excel_path()
    errors: list[str] = []
    if not path.is_file():
        errors.append(f"Excel file not found: {path}")
        return [], errors

    try:
        from openpyxl import load_workbook
    except ImportError:
        errors.append("openpyxl is required to load the control library Excel file")
        return [], errors

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        errors.append(f"Failed to open Excel workbook: {exc}")
        return [], errors

    sheet = None
    for name in workbook.sheetnames:
        if "consolidated" in name.lower() or "query" in name.lower() or "control" in name.lower():
            sheet = workbook[name]
            break
    if sheet is None:
        sheet = workbook[workbook.sheetnames[0]]

    rows_iter = sheet.iter_rows(values_only=True)
    header_row: tuple | None = None
    header_index = 0
    buffered: list[tuple] = []

    for i, row in enumerate(rows_iter):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        headers = [_normalize_header(cell) for cell in row]
        mapping = _resolve_header_index(list(row))
        matched = sum(1 for v in mapping.values() if v is not None)
        if matched >= 3:
            header_row = row
            header_index = i
            break
        if i < 10:
            buffered.append(row)

    if header_row is None:
        for row in buffered:
            mapping = _resolve_header_index(list(row))
            if sum(1 for v in mapping.values() if v is not None) >= 2:
                header_row = row
                break

    if header_row is None:
        errors.append("Could not detect header row in Excel — expected Control ID, Control Name, Query columns")
        workbook.close()
        return [], errors

    col_map = _resolve_header_index(list(header_row))
    if col_map["control_id"] is None and col_map["control_name"] is None:
        errors.append("Excel header row missing Control ID and Control Name columns")
        workbook.close()
        return [], errors

    controls: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    row_num = 0

    for row in sheet.iter_rows(values_only=True):
        row_num += 1
        if row_num <= header_index + 1:
            continue
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        control_id = _cell_value(row, col_map["control_id"])
        control_name = _cell_value(row, col_map["control_name"])
        framework_raw = _cell_value(row, col_map["framework_coverage"])
        query = _cell_value(row, col_map["query"])
        description = _cell_value(row, col_map["description"])
        evidence_type = _cell_value(row, col_map["evidence_type"])

        if not control_id and not control_name:
            continue
        if not control_id:
            control_id = control_name
        if control_id in seen_ids:
            errors.append(f"Duplicate Control ID skipped: {control_id}")
            continue
        seen_ids.add(control_id)

        frameworks = _parse_frameworks(framework_raw)
        predefined = bool(query)
        technology = detect_technology(query) if predefined else ""

        record = {
            "control_id": control_id,
            "control_name": control_name or control_id,
            "framework_coverage": ", ".join(frameworks) if frameworks else framework_raw,
            "frameworks": frameworks,
            "query": query,
            "description": description,
            "evidence_type": evidence_type,
            "predefined": predefined,
            "technology": technology,
            "status": "",
            "last_execution": "Not Executed",
        }
        capability = assess_execution_capability(record)
        record["status"] = capability["status"]
        record["executable"] = capability["executable"]
        record["capability_reason"] = capability["reason"]
        controls.append(record)

    workbook.close()
    if not controls:
        errors.append("No control rows loaded from Excel")
    return controls, errors


def _merge_supplementary_controls(controls: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Append code-defined supplementary DB controls (additive; Excel wins).

    Builds records with the SAME schema as ``_load_from_excel`` so the rest of the
    engine treats them identically. Existing control_ids from Excel are never
    overridden.
    """
    try:
        from modules.operations.engines.supplementary_query_catalog import supplementary_controls
    except Exception:  # noqa: BLE001 - optional, never break the primary load
        return controls

    existing_ids = {c.get("control_id") for c in controls}
    for entry in supplementary_controls():
        cid = entry.get("control_id")
        if not cid or cid in existing_ids:
            continue
        query = (entry.get("query") or "").strip()
        frameworks = _parse_frameworks(entry.get("framework_coverage") or "")
        # Prefer the explicitly-declared technology; fall back to text detection.
        technology = entry.get("technology") or detect_technology(query)
        record = {
            "control_id": cid,
            "control_name": entry.get("control_name") or cid,
            "framework_coverage": ", ".join(frameworks) if frameworks else (entry.get("framework_coverage") or ""),
            "frameworks": frameworks,
            "query": query,
            "description": entry.get("description") or "",
            "evidence_type": entry.get("evidence_type") or "",
            "predefined": bool(query),
            "technology": technology,
            # Optional metadata carried through from the catalog (used by UI/tests).
            "category": entry.get("category") or "",
            "status": "",
            "last_execution": "Not Executed",
            "source": "supplementary",
        }
        capability = assess_execution_capability(record)
        record["status"] = capability["status"]
        record["executable"] = capability["executable"]
        record["capability_reason"] = capability["reason"]
        controls.append(record)
        existing_ids.add(cid)
    return controls


def load_predefined_queries(*, force: bool = False) -> dict[str, Any]:
    """Load controls from Excel (idempotent)."""
    global _controls, _validation_report, _loaded, _errors_found, _errors_fixed

    if _loaded and not force:
        return _validation_report

    _errors_found = []
    _errors_fixed = []
    controls, load_errors = _load_from_excel()
    _errors_found.extend(load_errors)

    # Merge supplementary (code-defined) DB queries. Additive only: Excel entries
    # always win on control_id collision, so the workbook remains the source of
    # truth and this never overrides it.
    controls = _merge_supplementary_controls(controls)

    for ctrl in controls:
        if ctrl.get("predefined") and not ctrl.get("frameworks"):
            _errors_found.append(f"Missing framework mapping for {ctrl['control_id']}")
        if ctrl.get("predefined") and ctrl.get("technology") == "Unknown":
            _errors_found.append(f"Unsupported technology for {ctrl['control_id']}")

    for ctrl in controls:
        ctrl["last_execution"] = _last_execution(ctrl["control_id"])

    _controls = controls
    predefined_count = sum(1 for c in controls if c.get("predefined"))
    manual_count = len(controls) - predefined_count
    frameworks_covered = sorted({fw for c in controls for fw in c.get("frameworks", [])})

    _validation_report = {
        "excel_loaded": bool(controls) and not load_errors,
        "excel_path": str(_excel_path()),
        "controls_loaded": len(controls),
        "predefined_controls": predefined_count,
        "manual_controls": manual_count,
        "queries_loaded": predefined_count,
        "frameworks_covered": frameworks_covered,
        "technology_rules_loaded": len(TECHNOLOGY_RULES) > 0,
        "connector_configuration_loaded": _connector_config_loaded(),
        "errors_found": list(_errors_found),
        "errors_fixed": list(_errors_fixed),
    }
    _loaded = True
    return _validation_report


def validate_startup() -> dict[str, Any]:
    """Startup validation — returns report and log lines for terminal output."""
    report = load_predefined_queries()
    lines = [
        f"Excel Loaded: {'Yes' if report['excel_loaded'] else 'No'} ({report['excel_path']})",
        f"Controls Loaded: {report['controls_loaded']}",
        f"Predefined Controls: {report['predefined_controls']}",
        f"Manual Controls: {report['manual_controls']}",
        f"Queries Loaded: {report['queries_loaded']}",
        f"Frameworks Covered: {len(report['frameworks_covered'])} — {', '.join(report['frameworks_covered'][:8])}{'…' if len(report['frameworks_covered']) > 8 else ''}",
        f"Technology Rules Loaded: {report['technology_rules_loaded']}",
        f"Connector Configuration Loaded: {report['connector_configuration_loaded']}",
        f"Errors Found: {len(report['errors_found'])}",
        f"Errors Fixed: {len(report['errors_fixed'])}",
    ]
    if report["errors_found"]:
        for err in report["errors_found"][:5]:
            lines.append(f"  — {err}")
        if len(report["errors_found"]) > 5:
            lines.append(f"  — … and {len(report['errors_found']) - 5} more")
    report["log_lines"] = lines
    return report


def get_all_controls() -> list[dict[str, Any]]:
    if not _loaded:
        load_predefined_queries()
    return list(_controls)


def _normalize_query_allowlist(query: str) -> str:
    q = re.sub(r"\s+", " ", query.strip().lower())
    if q and not q.endswith(";"):
        q += ";"
    return q


def is_live_execution_enabled(control: dict[str, Any]) -> bool:
    # A control is live-executable only when it is in the curated live set AND
    # genuinely executable (implemented connector + dependency present). This
    # prevents enabling "Run Query" for controls whose status is not Ready
    # (e.g. an Unknown-technology control that was listed in LIVE_CONTROL_IDS).
    return (
        bool(control.get("predefined"))
        and control.get("control_id") in LIVE_CONTROL_IDS
        and assess_execution_capability(control)["executable"]
    )


def is_postgresql_execution_enabled(control: dict[str, Any]) -> bool:
    if not control.get("predefined"):
        return False
    if control.get("technology") != "PostgreSQL":
        return False
    normalized = _normalize_query_allowlist(control.get("query") or "")
    return normalized in ALLOWED_POSTGRESQL_QUERIES


def _refresh_control_last_execution(control_id: str) -> None:
    last = _last_execution(control_id)
    for ctrl in _controls:
        if ctrl["control_id"] == control_id:
            ctrl["last_execution"] = last
            break


def get_control_by_id(control_id: str) -> dict[str, Any] | None:
    for ctrl in get_all_controls():
        if ctrl["control_id"] == control_id:
            enriched = dict(ctrl)
            history = get_execution_history_for_control(control_id)
            enriched["execution_history"] = history
            enriched["latest_result"] = get_latest_evidence_for_control(control_id)
            enriched["latest_execution"] = history[0] if history else None
            return enriched
    return None


def _runtime_suggested_action(failure_point: str) -> str:
    return {
        "Connector connection": "Confirm the target service is running and reachable, then re-run the query.",
        "Driver / dependency": "Install the required driver in the execution environment, then re-run the query.",
        "Execution gate / allow-list": "Enable this control for live execution (target / query allow-list), then re-run.",
        "Query execution": "Review the query and target permissions, then re-run.",
    }.get(failure_point, "Review the connector configuration and re-run the query.")


def derive_runtime_state(control: dict[str, Any] | None) -> dict[str, Any]:
    """Single source of truth for the UI execution banner.

    The execution status is derived ONLY from the durable runtime signals
    (``latest_execution`` / ``latest_result`` / ``execution_history``) — never
    from a (possibly stale) URL ``notice`` parameter. Rule: when an evidence
    record exists the control's execution is treated as SUCCESS regardless of
    any stale failure notice, so SUCCESS always overrides a stale banner and the
    two can never be shown together.
    """
    control = control or {}
    le = control.get("latest_execution") or {}
    lr = control.get("latest_result") or {}
    history = control.get("execution_history") or []

    has_evidence = bool(lr.get("evidence_id"))
    succeeded = str(le.get("status", "")).lower() == "success" or has_evidence
    executed = bool(le) or has_evidence or bool(history)

    if succeeded:
        status = "SUCCESS"
    elif le:
        status = "FAILED"
    else:
        status = "NOT_EXECUTED"

    err = le.get("error_message") or ""
    low = err.lower()
    if not err or succeeded:
        failure_point = ""
    elif any(k in low for k in (
        "reachable", "could not connect", "connection refused", "timed out",
        "timeout", "no route", "unavailable", "not connected",
    )):
        failure_point = "Connector connection"
    elif "not installed" in low or "driver" in low:
        failure_point = "Driver / dependency"
    elif any(k in low for k in ("not enabled", "allow-list", "unsupported")):
        failure_point = "Execution gate / allow-list"
    else:
        failure_point = "Query execution"

    return {
        "status": status,
        "executed": executed,
        "success": succeeded,
        "failed": status == "FAILED",
        "rows_returned": le.get("rows_returned"),
        "duration_ms": le.get("duration_ms"),
        "evidence_id": lr.get("evidence_id") or "",
        "timestamp": le.get("execution_time") or "",
        "user": le.get("user") or "",
        "connector": control.get("technology") or "",
        "error_message": "" if succeeded else err,
        "failure_point": failure_point,
        "suggested_action": _runtime_suggested_action(failure_point) if failure_point else "",
        # When the run succeeded, the route must drop the stale notice so a
        # success banner and a failure/connector/dependency banner never co-exist.
        "suppress_notice": succeeded,
    }


def get_framework_filter_options() -> list[str]:
    frameworks = {"All Frameworks"}
    for ctrl in get_all_controls():
        for fw in ctrl.get("frameworks", []):
            frameworks.add(fw)
    return ["All Frameworks"] + sorted(frameworks - {"All Frameworks"})


def get_technology_filter_options() -> list[str]:
    """Distinct technologies present in the catalog (for the Technology filter)."""
    techs = {
        (ctrl.get("technology") or "").strip()
        for ctrl in get_all_controls()
        if (ctrl.get("technology") or "").strip()
    }
    return ["All Technologies"] + sorted(techs)


def filter_controls(
    *,
    search: str = "",
    framework: str = "All Frameworks",
    technology: str = "All Technologies",
    predefined_only: bool = False,
    sort_by: str = "control_id",
    sort_dir: str = "asc",
) -> list[dict[str, Any]]:
    rows = get_all_controls()
    q = search.strip().lower()

    if framework and framework != "All Frameworks":
        rows = [r for r in rows if framework in r.get("frameworks", []) or framework in (r.get("framework_coverage") or "")]

    if technology and technology != "All Technologies":
        rows = [r for r in rows if (r.get("technology") or "") == technology]

    if predefined_only:
        rows = [r for r in rows if r.get("predefined")]

    if q:
        rows = [
            r
            for r in rows
            if q in r["control_id"].lower()
            or q in r["control_name"].lower()
            or q in (r.get("query") or "").lower()
            or q in (r.get("framework_coverage") or "").lower()
            or q in (r.get("technology") or "").lower()
        ]

    reverse = sort_dir.lower() == "desc"
    sort_key = sort_by if sort_by in ("control_id", "control_name", "technology", "status", "last_execution") else "control_id"
    rows.sort(key=lambda r: (r.get(sort_key) or "").lower(), reverse=reverse)
    return rows


def get_predefined_queries_dashboard(
    *,
    search: str = "",
    framework: str = "All Frameworks",
    technology: str = "All Technologies",
    page: int = 1,
    per_page: int = 10,
    sort_by: str = "control_id",
    sort_dir: str = "asc",
    predefined_only: bool = False,
) -> dict[str, Any]:
    """Build dashboard payload for the Predefined Queries page."""
    report = load_predefined_queries()
    filtered = filter_controls(
        search=search,
        framework=framework,
        technology=technology,
        predefined_only=predefined_only,
        sort_by=sort_by,
        sort_dir=sort_dir,
    )
    page_data = paginate(filtered, page=page, per_page=per_page)
    for item in page_data["items"]:
        item["live_execution_enabled"] = is_live_execution_enabled(item)

    predefined_rows = [r for r in get_all_controls() if r.get("predefined")]
    manual_rows = [r for r in get_all_controls() if not r.get("predefined")]
    unsupported = [r for r in predefined_rows if r.get("technology") == "Unknown"]

    return {
        "validation": report,
        "rows": page_data["items"],
        "pagination": page_data,
        "framework_options": get_framework_filter_options(),
        "technology_options": get_technology_filter_options(),
        "search": search,
        "framework_filter": framework,
        "technology_filter": technology,
        "sort_by": sort_by,
        "sort_dir": sort_dir,
        "predefined_only": predefined_only,
        "kpis": [
            {"label": "Total Controls", "value": report["controls_loaded"], "tone": "primary"},
            {"label": "Predefined Queries", "value": report["predefined_controls"], "tone": "success"},
            {"label": "Manual Controls", "value": report["manual_controls"], "tone": "secondary"},
            {"label": "Frameworks Covered", "value": len(report["frameworks_covered"]), "tone": "info"},
            {"label": "Unsupported Tech", "value": len(unsupported), "tone": "warning"},
        ],
        "all_predefined_count": len(predefined_rows),
        "manual_count": len(manual_rows),
        "unsupported_count": len(unsupported),
        "execution_enabled": True,
        "execution_history_rows": get_execution_audit_log(limit=100),
    }


def _pq_drill_body(title: str, columns: list[str], rows: list[dict[str, Any]], *, empty_note: str) -> dict[str, Any]:
    body: dict[str, Any] = {
        "ok": True,
        "title": f"{title} — Predefined Queries",
        "columns": columns,
        "rows": rows,
        "row_count": len(rows),
    }
    if not rows:
        # Honest empty-state: no fabricated, padded, or unrelated records.
        body["note"] = empty_note
    return body


def drill_predefined_query_kpi(metric: str, role: str = "owner") -> dict[str, Any]:
    """Authoritative, traceable drilldown for the Predefined Queries KPI cards.

    Each KPI returns ONLY rows directly related to that KPI, sourced from the
    loaded control library. When a KPI count is zero an honest empty-state is
    returned (explanatory note, no rows) — never fabricated, padded, or
    unrelated records. This is the single source of truth for KPI drilldowns on
    the Predefined Queries page.
    """
    load_predefined_queries()
    controls = get_all_controls()
    m = (metric or "").strip().lower().replace("-", "_").replace(" ", "_")

    control_columns = ["control", "control_name", "framework", "technology", "status"]

    def _control_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [
            {
                "control": c.get("control_id", "—"),
                "control_name": c.get("control_name", "—"),
                "framework": c.get("framework_coverage") or "—",
                "technology": c.get("technology") or "—",
                "status": c.get("status") or "—",
            }
            for c in rows
        ]

    if m in ("total_controls", "total", "controls"):
        return _pq_drill_body(
            "Total Controls", control_columns, _control_rows(controls),
            empty_note="No controls are loaded from the control library.",
        )

    if m in ("predefined_queries", "predefined", "queries", "ready_queries", "ready"):
        rows = [c for c in controls if c.get("predefined")]
        return _pq_drill_body(
            "Predefined Queries", control_columns, _control_rows(rows),
            empty_note="No predefined queries are loaded from the control library.",
        )

    if m in ("manual_controls", "manual"):
        rows = [c for c in controls if not c.get("predefined")]
        return _pq_drill_body(
            "Manual Controls", control_columns, _control_rows(rows),
            empty_note="No manual controls — every loaded control has a predefined query.",
        )

    if m in ("frameworks_covered", "frameworks", "framework_coverage"):
        counts: dict[str, dict[str, int]] = {}
        for c in controls:
            for fw in c.get("frameworks", []):
                slot = counts.setdefault(fw, {"controls": 0, "predefined": 0, "manual": 0})
                slot["controls"] += 1
                if c.get("predefined"):
                    slot["predefined"] += 1
                else:
                    slot["manual"] += 1
        rows = [
            {"framework": fw, "controls": v["controls"], "predefined": v["predefined"], "manual": v["manual"]}
            for fw, v in sorted(counts.items())
        ]
        return _pq_drill_body(
            "Frameworks Covered", ["framework", "controls", "predefined", "manual"], rows,
            empty_note="No frameworks are mapped to the loaded controls.",
        )

    if m in ("unsupported_tech", "unsupported_technology", "unsupported"):
        rows = [
            {
                "control": c.get("control_id", "—"),
                "control_name": c.get("control_name", "—"),
                "framework": c.get("framework_coverage") or "—",
                "query_excerpt": (c.get("query") or "—")[:60] + ("…" if len(c.get("query") or "") > 60 else ""),
                "reason": c.get("capability_reason")
                or "Query text did not match any known technology pattern, so no connector can run it.",
            }
            for c in controls
            if c.get("predefined") and c.get("technology") == "Unknown"
        ]
        return _pq_drill_body(
            "Unsupported Technology",
            ["control", "control_name", "framework", "query_excerpt", "reason"], rows,
            empty_note="No unsupported technologies — every predefined query maps to a known connector.",
        )

    # Unknown metric → return the full catalog rather than unrelated/fabricated data.
    return _pq_drill_body(
        (metric or "Predefined Queries").replace("_", " ").title(),
        control_columns, _control_rows(controls),
        empty_note="No controls are loaded from the control library.",
    )


def prepare_execution(control_id: str, user: str) -> dict[str, Any]:
    """
    Prepare execution context (interfaces only — does not execute).
    Returns structured response for UI; never raises to caller.
    """
    control = get_control_by_id(control_id)
    if not control:
        return {"ok": False, "error": "Control not found", "error_type": "missing_control"}

    if not control.get("query"):
        return {"ok": False, "error": "Missing query for this control", "error_type": "missing_query"}

    technology = control.get("technology") or ""
    if not technology or technology == "Unknown":
        return {
            "ok": False,
            "error": "Technology could not be determined for this query",
            "error_type": "unsupported_technology",
        }

    if not control.get("frameworks") and not control.get("framework_coverage"):
        return {
            "ok": False,
            "error": "Framework mapping missing for this control",
            "error_type": "missing_framework",
        }

    from modules.operations.engines.query_connectors import connector_for_technology

    connector = connector_for_technology(technology)
    if connector is None:
        return {
            "ok": False,
            "error": f"No connector available for technology: {technology}",
            "error_type": "unsupported_technology",
        }

    execution_enabled = is_live_execution_enabled(control)
    return {
        "ok": True,
        "execution_enabled": execution_enabled,
        "message": "Live demo execution enabled" if execution_enabled else "Execution not enabled for this control",
        "control_id": control_id,
        "technology": technology,
        "connector": type(connector).__name__,
        "user": user,
    }


def run_postgresql_query(control_id: str, user: str) -> dict[str, Any]:
    """Execute a predefined PostgreSQL query from the Excel catalog."""
    control = get_control_by_id(control_id)
    if not control:
        return {"ok": False, "error": "Control not found", "error_type": "missing_control"}

    query = (control.get("query") or "").strip()
    if not query:
        return {"ok": False, "error": "Missing query for this control", "error_type": "missing_query"}

    if control.get("technology") != "PostgreSQL":
        return {
            "ok": False,
            "error": "Live execution is only enabled for PostgreSQL controls",
            "error_type": "unsupported_technology",
        }

    if _normalize_query_allowlist(query) not in ALLOWED_POSTGRESQL_QUERIES:
        return {
            "ok": False,
            "error": "This PostgreSQL query is not enabled for live demo execution",
            "error_type": "unsupported_query",
        }

    try:
        from modules.operations.engines.postgresql_connector import PostgreSQLConnector, get_postgresql_config
    except ImportError as exc:
        # The PostgreSQL driver (psycopg2) is not available in this environment.
        # Return a structured, demo-friendly error rather than letting a raw
        # ModuleNotFoundError bubble up into a 500 / stack trace.
        missing = getattr(exc, "name", "") or "psycopg2"
        return {
            "ok": False,
            "error": "PostgreSQL connector unavailable",
            "error_type": "connector_unavailable",
            "reason": f"Required driver '{missing}' is not installed in this environment.",
            "action": "Install psycopg2-binary or configure the PostgreSQL connector, then retry.",
        }

    connector = PostgreSQLConnector(**get_postgresql_config())
    framework_coverage = control.get("framework_coverage") or ""

    if not connector.connect():
        record_execution_audit(
            control_id,
            user,
            "PostgreSQL",
            "Failed",
            error_message=connector._last_error or "Connection failed",
            framework_coverage=framework_coverage,
            query=query,
        )
        _refresh_control_last_execution(control_id)
        return {
            "ok": False,
            "error": connector._last_error or "Could not connect to PostgreSQL",
            "error_type": "connection_failure",
        }

    try:
        result = connector.execute(query)
    finally:
        connector.disconnect()

    rows_returned = int(result.metadata.get("rows_returned", 0)) if result.metadata else 0

    if not result.success:
        record_execution_audit(
            control_id,
            user,
            "PostgreSQL",
            "Failed",
            duration_ms=result.duration_ms,
            error_message=result.error_message,
            framework_coverage=framework_coverage,
            query=query,
            result=result.output,
            rows_returned=rows_returned,
        )
        _refresh_control_last_execution(control_id)
        return {
            "ok": False,
            "error": result.error_message,
            "error_type": result.metadata.get("error_type", "query_failure") if result.metadata else "query_failure",
        }

    record_execution_audit(
        control_id,
        user,
        "PostgreSQL",
        "Success",
        duration_ms=result.duration_ms,
        framework_coverage=framework_coverage,
        query=query,
        result=result.output,
        rows_returned=rows_returned,
    )

    evidence = prepare_evidence_record(
        control_id=control_id,
        result=result.output,
        user=user,
        framework_coverage=framework_coverage,
    )
    primary_fw = control.get("frameworks", [""])[0] if control.get("frameworks") else ""
    register_with_evidence_repository(evidence, framework=primary_fw)
    _refresh_control_last_execution(control_id)

    from modules.shared.services.audit_trail import log_event

    log_event(
        "Predefined Query Executed",
        user,
        primary_fw,
        control_id,
        f"PostgreSQL — {rows_returned} rows — {result.duration_ms}ms",
    )

    return {
        "ok": True,
        "message": f"Query executed successfully — {rows_returned} row(s) returned in {result.duration_ms}ms",
        "control_id": control_id,
        "query": query,
        "status": "Success",
        "rows_returned": rows_returned,
        "output": result.output,
        "duration_ms": result.duration_ms,
        "evidence_id": evidence.evidence_id,
    }


def run_yugabyte_query(control_id: str, user: str) -> dict[str, Any]:
    """Execute a predefined YugabyteDB (YSQL) query from the catalog.

    Reuses the PostgreSQL-wire connector; enforces the Yugabyte allow-list and the
    shared audit/evidence completion path used by other connectors.
    """
    control = get_control_by_id(control_id)
    if not control:
        return {"ok": False, "error": "Control not found", "error_type": "missing_control"}

    query = (control.get("query") or "").strip()
    if not query:
        return {"ok": False, "error": "Missing query for this control", "error_type": "missing_query"}

    if control.get("technology") != "YugabyteDB":
        return {"ok": False, "error": "This control is not a YugabyteDB control",
                "error_type": "unsupported_technology"}

    if _normalize_query_allowlist(query) not in ALLOWED_YUGABYTE_QUERIES:
        return {"ok": False, "error": "This YugabyteDB query is not enabled for live execution",
                "error_type": "unsupported_query"}

    try:
        from modules.operations.engines.yugabyte_connector import YugabyteConnector, get_yugabyte_config
    except ImportError as exc:
        missing = getattr(exc, "name", "") or "psycopg2"
        return {
            "ok": False,
            "error": "YugabyteDB connector unavailable",
            "error_type": "connector_unavailable",
            "reason": f"Required driver '{missing}' is not installed in this environment.",
            "action": "Install psycopg2-binary, then retry.",
        }

    connector = YugabyteConnector(**get_yugabyte_config())
    return _run_connector_query(control, user, "YugabyteDB", query, connector)


def run_mysql_query(control_id: str, user: str) -> dict[str, Any]:
    """Execute a predefined Aurora MySQL query from the catalog.

    Uses the PyMySQL connector; enforces the MySQL allow-list and the shared
    audit/evidence completion path used by other connectors.
    """
    control = get_control_by_id(control_id)
    if not control:
        return {"ok": False, "error": "Control not found", "error_type": "missing_control"}

    query = (control.get("query") or "").strip()
    if not query:
        return {"ok": False, "error": "Missing query for this control", "error_type": "missing_query"}

    if control.get("technology") != "Aurora MySQL":
        return {"ok": False, "error": "This control is not an Aurora MySQL control",
                "error_type": "unsupported_technology"}

    if _normalize_query_allowlist(query) not in ALLOWED_MYSQL_QUERIES:
        return {"ok": False, "error": "This Aurora MySQL query is not enabled for live execution",
                "error_type": "unsupported_query"}

    try:
        from modules.operations.engines.mysql_connector import MySQLConnector, get_mysql_config
    except ImportError as exc:
        missing = getattr(exc, "name", "") or "pymysql"
        return {
            "ok": False,
            "error": "Aurora MySQL connector unavailable",
            "error_type": "connector_unavailable",
            "reason": f"Required driver '{missing}' is not installed in this environment.",
            "action": "Install PyMySQL, then retry.",
        }

    connector = MySQLConnector(**get_mysql_config())
    return _run_connector_query(control, user, "Aurora MySQL", query, connector)


def run_oracle_query(control_id: str, user: str) -> dict[str, Any]:
    """Execute a predefined Oracle query from the catalog (python-oracledb)."""
    control = get_control_by_id(control_id)
    if not control:
        return {"ok": False, "error": "Control not found", "error_type": "missing_control"}

    query = (control.get("query") or "").strip()
    if not query:
        return {"ok": False, "error": "Missing query for this control", "error_type": "missing_query"}

    if control.get("technology") != "Oracle":
        return {"ok": False, "error": "This control is not an Oracle control",
                "error_type": "unsupported_technology"}

    if _normalize_query_allowlist(query) not in ALLOWED_ORACLE_QUERIES:
        return {"ok": False, "error": "This Oracle query is not enabled for live execution",
                "error_type": "unsupported_query"}

    try:
        from modules.operations.engines.oracle_connector import OracleConnector, get_oracle_config
    except ImportError as exc:
        missing = getattr(exc, "name", "") or "oracledb"
        return {
            "ok": False,
            "error": "Oracle connector unavailable",
            "error_type": "connector_unavailable",
            "reason": f"Required driver '{missing}' is not installed in this environment.",
            "action": "Install python-oracledb, then retry.",
        }

    connector = OracleConnector(**get_oracle_config())
    return _run_connector_query(control, user, "Oracle", query, connector)


def run_sqlserver_query(control_id: str, user: str) -> dict[str, Any]:
    """Execute a predefined SQL Server query from the catalog (pyodbc)."""
    control = get_control_by_id(control_id)
    if not control:
        return {"ok": False, "error": "Control not found", "error_type": "missing_control"}
    query = (control.get("query") or "").strip()
    if not query:
        return {"ok": False, "error": "Missing query for this control", "error_type": "missing_query"}
    if control.get("technology") != "SQL Server":
        return {"ok": False, "error": "This control is not a SQL Server control",
                "error_type": "unsupported_technology"}
    if _norm_sql(query) not in ALLOWED_SQLSERVER_QUERIES:
        return {"ok": False, "error": "This SQL Server query is not enabled for live execution",
                "error_type": "unsupported_query"}
    try:
        from modules.operations.engines.sqlserver_connector import SQLServerConnector, get_sqlserver_config
    except ImportError as exc:
        missing = getattr(exc, "name", "") or "pyodbc"
        return {
            "ok": False, "error": "SQL Server connector unavailable",
            "error_type": "connector_unavailable",
            "reason": f"Required driver '{missing}' is not installed in this environment.",
            "action": "Install pyodbc and an ODBC driver, then retry.",
        }
    connector = SQLServerConnector(**get_sqlserver_config())
    return _run_connector_query(control, user, "SQL Server", query, connector)


def run_mongodb_query(control_id: str, user: str) -> dict[str, Any]:
    """Execute a predefined MongoDB admin command from the catalog (pymongo)."""
    control = get_control_by_id(control_id)
    if not control:
        return {"ok": False, "error": "Control not found", "error_type": "missing_control"}
    command = (control.get("query") or "").strip()
    if not command:
        return {"ok": False, "error": "Missing command for this control", "error_type": "missing_query"}
    if control.get("technology") != "MongoDB":
        return {"ok": False, "error": "This control is not a MongoDB control",
                "error_type": "unsupported_technology"}
    if command not in ALLOWED_MONGODB_COMMANDS:
        return {"ok": False, "error": "This MongoDB command is not enabled for live execution",
                "error_type": "unsupported_query"}
    try:
        from modules.operations.engines.mongodb_connector import MongoDBConnector, get_mongodb_config
    except ImportError as exc:
        missing = getattr(exc, "name", "") or "pymongo"
        return {
            "ok": False, "error": "MongoDB connector unavailable",
            "error_type": "connector_unavailable",
            "reason": f"Required driver '{missing}' is not installed in this environment.",
            "action": "Install pymongo, then retry.",
        }
    connector = MongoDBConnector(**get_mongodb_config())
    return _run_connector_query(control, user, "MongoDB", command, connector)


def run_shell_control(control_id: str, user: str) -> dict[str, Any]:
    """Execute a predefined shell/CLI-command control.

    Covers Linux / NGINX / RHEL 8.x / 9.x / Redis / Apache HTTPD / Tomcat via the
    docker-exec Linux connector (Redis via redis-cli in its container), and
    Kubernetes / OpenShift via local kubectl / oc subprocesses. The command comes
    exclusively from the curated code-defined catalog (never user input) and the
    control must be in the shell allow-list.
    """
    control = get_control_by_id(control_id)
    if not control:
        return {"ok": False, "error": "Control not found", "error_type": "missing_control"}

    technology = control.get("technology") or ""
    if technology not in _SHELL_TECHNOLOGIES:
        return {"ok": False, "error": f"This control is not a shell control ({technology})",
                "error_type": "unsupported_technology"}

    command = (control.get("query") or "").strip()
    if not command:
        return {"ok": False, "error": "Missing command for this control", "error_type": "missing_query"}

    # Only curated catalog commands (by control id) may run.
    if control_id not in _SHELL_CONTROL_IDS and control_id not in LINUX_CONTROL_COMMANDS_ALL:
        return {"ok": False, "error": "This control is not enabled for live execution",
                "error_type": "unsupported_query"}

    from modules.operations.engines.linux_connector import (
        LinuxConnector,
        get_apache_config,
        get_linux_config,
        get_nginx_config,
        get_rhel_config,
        get_tomcat_config,
    )

    if technology == "NGINX":
        connector = LinuxConnector(**get_nginx_config())
    elif technology == RHEL8_TECH:
        connector = LinuxConnector(**get_rhel_config(8))
    elif technology == RHEL9_TECH:
        connector = LinuxConnector(**get_rhel_config(9))
    elif technology == "Apache HTTPD":
        connector = LinuxConnector(**get_apache_config())
    elif technology == "Tomcat":
        connector = LinuxConnector(**get_tomcat_config())
    elif technology == "Redis":
        from modules.operations.engines.redis_connector import RedisConnector, get_redis_config

        connector = RedisConnector(**get_redis_config())
    elif technology == "Kubernetes":
        from modules.operations.engines.kubernetes_connector import (
            KubernetesConnector,
            get_kubernetes_config,
        )

        connector = KubernetesConnector(**get_kubernetes_config())
    elif technology == "OpenShift":
        from modules.operations.engines.kubernetes_connector import (
            OpenShiftConnector,
            get_openshift_config,
        )

        connector = OpenShiftConnector(**get_openshift_config())
    else:  # Linux
        connector = LinuxConnector(**get_linux_config())
    return _run_connector_query(control, user, technology, command, connector)


# Set of control_ids whose Linux command is defined either in the legacy
# LINUX_CONTROL_COMMANDS map or the supplementary shell catalog.
def _linux_control_commands_all() -> frozenset[str]:
    try:
        from modules.operations.engines.linux_connector import LINUX_CONTROL_COMMANDS

        return frozenset(LINUX_CONTROL_COMMANDS.keys())
    except Exception:  # noqa: BLE001
        return frozenset()


LINUX_CONTROL_COMMANDS_ALL: frozenset[str] = _linux_control_commands_all()


def _run_connector_query(control: dict[str, Any], user: str, technology: str, query: str, connector) -> dict[str, Any]:
    from modules.operations.engines.connector_common import complete_connector_execution
    from modules.operations.engines.query_connectors import ConnectorResult

    if not connector.connect():
        err = getattr(connector, "_last_error", "") or "Connection failed"
        return complete_connector_execution(
            control,
            user,
            technology,
            query,
            ConnectorResult(success=False, error_message=err),
            connect_error=err,
        )
    try:
        result = connector.execute(query)
    finally:
        connector.disconnect()
    return complete_connector_execution(control, user, technology, query, result)


def run_aerospike_query(control_id: str, user: str) -> dict[str, Any]:
    """Execute an Aerospike asinfo/asadm control via the Aerospike connector.

    Attempts live execution via ``docker exec ecs-aerospike ...``. If the
    container / tools are unavailable AND demo mode is on, returns deterministic
    synthetic output (never a live call to an unconfigured target, never a crash).
    Outside demo mode an unavailable target returns a truthful connection error.
    """
    control = get_control_by_id(control_id)
    if not control:
        return {"ok": False, "error": "Control not found", "error_type": "missing_control"}

    technology = control.get("technology") or ""
    if technology != "Aerospike":
        return {"ok": False, "error": f"This control is not an Aerospike control ({technology})",
                "error_type": "unsupported_technology"}

    command = (control.get("query") or "").strip()
    if not command:
        return {"ok": False, "error": "Missing command for this control", "error_type": "missing_query"}

    # Only curated catalog commands (by control id) may run.
    if control_id not in _SHELL_CONTROL_IDS:
        return {"ok": False, "error": "This control is not enabled for live execution",
                "error_type": "unsupported_query"}

    from modules.operations.engines.aerospike_connector import (
        AerospikeConnector,
        _demo_mode,
        _resolve_namespace,
        demo_output_for,
        get_aerospike_config,
    )
    from modules.operations.engines.connector_common import complete_connector_execution
    from modules.operations.engines.query_connectors import ConnectorResult

    cfg = get_aerospike_config()

    # Demo mode: return deterministic synthetic output WITHOUT touching a live
    # target (fast, offline, no docker dependency). This is the ECS demo-first
    # pattern — live execution is reserved for an explicitly-configured UAT node.
    if _demo_mode():
        resolved = _resolve_namespace(command, cfg.get("namespace") or "test")
        output = demo_output_for(control_id, resolved)
        result = ConnectorResult(
            success=True, output=output, duration_ms=1,
            metadata={"rows_returned": len(output.splitlines()), "mode": "demo"},
        )
        return complete_connector_execution(control, user, technology, command, result)

    # Live path (non-demo): run asinfo/asadm against the configured node.
    connector = AerospikeConnector(**cfg)
    if connector.connect():
        try:
            result = connector.execute(command)
        finally:
            connector.disconnect()
        return complete_connector_execution(control, user, technology, command, result)

    err = getattr(connector, "_last_error", "") or "Aerospike target not reachable"
    return {"ok": False, "error": err, "error_type": "connection_error"}


def run_predefined_query(control_id: str, user: str) -> dict[str, Any]:
    """Dispatch live execution by control ID and technology."""
    control = get_control_by_id(control_id)
    if not control:
        return {"ok": False, "error": "Control not found", "error_type": "missing_control"}
    if not is_live_execution_enabled(control):
        return {
            "ok": False,
            "error": "Live execution is not enabled for this control",
            "error_type": "unsupported_control",
        }

    technology = control.get("technology") or ""
    if technology == "PostgreSQL":
        return run_postgresql_query(control_id, user)

    if technology == "YugabyteDB":
        return run_yugabyte_query(control_id, user)

    if technology == "Aurora MySQL":
        return run_mysql_query(control_id, user)

    if technology == "Oracle":
        return run_oracle_query(control_id, user)

    if technology == "SQL Server":
        return run_sqlserver_query(control_id, user)

    if technology == "MongoDB":
        return run_mongodb_query(control_id, user)

    query = (control.get("query") or "").strip()

    # Aerospike runs asinfo/asadm CLI commands against the node. Handled before the
    # generic shell path (which is Linux-docker-exec specific). Falls back to
    # deterministic demo output in demo mode when the tools are absent.
    if technology == "Aerospike":
        return run_aerospike_query(control_id, user)

    # NGINX / RHEL / Redis / Apache / Tomcat / Kubernetes / OpenShift run curated
    # shell or CLI commands via the shared command connectors (reuse, not
    # duplication).
    if technology in _SHELL_TECHNOLOGIES and technology != "Linux":
        return run_shell_control(control_id, user)

    if technology == "Linux":
        from modules.operations.engines.linux_connector import (
            LINUX_CONTROL_COMMANDS,
            LinuxConnector,
            get_linux_config,
        )

        command = LINUX_CONTROL_COMMANDS.get(control_id, query)
        connector = LinuxConnector(**get_linux_config())
        return _run_connector_query(control, user, technology, command, connector)

    if technology == "SonarQube":
        from modules.operations.engines.sonarqube_connector import SonarQubeConnector, get_sonarqube_config

        mode = SONAR_CONTROL_MODES.get(control_id, query)
        connector = SonarQubeConnector(**get_sonarqube_config())
        return _run_connector_query(control, user, technology, mode, connector)

    if technology == "Trivy":
        from modules.operations.engines.trivy_connector import TrivyConnector, get_trivy_config

        connector = TrivyConnector(**get_trivy_config())
        return _run_connector_query(control, user, technology, query or "trivy image", connector)

    if technology == "GitLeaks":
        from modules.operations.engines.gitleaks_connector import GitLeaksConnector, get_gitleaks_config

        connector = GitLeaksConnector(**get_gitleaks_config())
        return _run_connector_query(control, user, technology, query or "gitleaks detect", connector)

    return {
        "ok": False,
        "error": f"Live execution is not enabled for technology: {technology}",
        "error_type": "unsupported_technology",
    }
