#!/usr/bin/env python3
"""Regenerate docs/use-cases/MD_Usecases_Implementation_Assessment.xlsx from repo state."""

from __future__ import annotations

from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from modules.operations.engines.common_controls_catalog import COMMON_CONTROLS

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs/use-cases/MD_Usecases_Implementation_Assessment.xlsx"

FW = "PCI DSS, DPSC, ISG Baseline, ISO 27001, VAPT, IS, CSITE, AppSec, ITPP"
PHASE_ORDER = {"Phase1": 1, "Phase2": 2, "Phase3": 3, "Phase4": 4}


def _exists(paths: str) -> str:
    ok = []
    for p in paths.replace(",", ";").split(";"):
        p = p.strip()
        if not p:
            continue
        if (ROOT / p).exists():
            ok.append(p)
    return "; ".join(ok)


USECASES = [
    {
        "name": "Manual evidence upload",
        "phase": "Phase1",
        "controls": "Evidence collection control, audit trail, access control",
        "frameworks": FW,
        "why": "Allows teams to submit audit evidence with traceability for regulatory and internal audits.",
        "status": "Implemented (MVP/demo)",
        "justification": "Single-file upload via evidence API and MVP upload routes. Content is hashed, named with enforce_naming(), registered in operations evidence_repository and mirrored to audit-intelligence repository when bridge succeeds.",
        "screen": "Administration → Bulk Evidence Upload → select framework/application → upload → repository/health views.",
        "sources": "modules/shared/routes/routes_mvp.py; modules/shared/services/evidence_api.py; modules/operations/engines/evidence_repository.py; modules/operations/templates/mvp_bulk_upload.html; tests/test_usecase_batch1_evidence_workflows.py",
        "design": "Route → upload service → naming/hash/metadata → dual evidence stack (MVP + audit-intelligence).",
        "remarks": "MVP/demo path is executable. Object-store persistence is opt-in via ECS evidence custody config.",
        "agenting": "No",
        "agenting_note": "Deterministic upload, hash and registration — no autonomous agent orchestration.",
    },
    {
        "name": "Bulk evidence upload",
        "phase": "Phase1",
        "controls": "Evidence collection control, audit trail, access control",
        "frameworks": FW,
        "why": "Supports mass onboarding of evidence artefacts during audit windows.",
        "status": "Implemented (MVP/demo)",
        "justification": "POST /mvp/upload/bulk handles multipart batches; each file flows through register_upload() with the same audit-repository bridge as manual upload.",
        "screen": "Administration → Bulk Evidence Upload → multi-select files → bulk POST → batch summary.",
        "sources": "modules/shared/routes/routes_mvp.py; modules/operations/engines/evidence_repository.py; modules/operations/templates/mvp_bulk_upload.html; tests/test_usecase_batch1_evidence_workflows.py",
        "design": "Multipart form → per-file register_upload → shared naming/hash pipeline.",
        "remarks": "Batch limits and virus scanning are deployment concerns — out of MVP implementation scope.",
        "agenting": "No",
        "agenting_note": "Batch ingest is scripted/form-driven, not agent-led.",
    },
    {
        "name": "Automated scheduled evidence pull",
        "phase": "Phase1",
        "controls": "Encryption at rest, encryption in transit, automated evidence collection",
        "frameworks": FW,
        "why": "Automates retrieval of configuration evidence required for recurring audits.",
        "status": "Partially implemented",
        "justification": "Scheduler UI, POST /mvp/scheduler/run, asset_scheduler planning, scheduler_module.run_scheduler_collection(), connector_executor wiring, and CommonControls/ folder discovery + collection (common_controls_collector) with tests (tests/test_scheduler_run_wiring.py, tests/test_common_controls_scheduler.py). Live connector fetch requires ECS_CONNECTOR_EXECUTION_ENABLED.",
        "screen": "Administration → Scheduler → Run Evidence Collection modal → staged progress → scheduler summary on reload.",
        "sources": "modules/operations/templates/mvp_scheduler.html; modules/operations/templates/partials/scheduler_modals.html; modules/operations/engines/scheduler_module.py; modules/operations/engines/common_controls_collector.py; CommonControls/; modules/audit_intelligence/services/asset_scheduler.py; modules/shared/routes/routes_mvp.py; tests/test_scheduler_run_wiring.py; tests/test_common_controls_scheduler.py",
        "design": "Scheduler request → plan jobs → connector execution (dry-run default) → evidence registration → redirect notice.",
        "remarks": "Assumption: demo operations dataset backs KPI tiles when connectors are not live. Progress UI is client-side staged animation before form POST.",
        "agenting": "No",
        "agenting_note": "Orchestrator executes planned connector jobs; no LLM/agent loop.",
    },
    {
        "name": "Metadata tagging and naming convention",
        "phase": "Phase1",
        "controls": "Evidence classification, control mapping",
        "frameworks": FW,
        "why": "Ensures evidence is discoverable and mapped to controls across frameworks.",
        "status": "Partially implemented",
        "justification": "enforce_naming() applied at ingest; naming preview/validate APIs exist; audit repository search supports tag/framework/technology filters.",
        "screen": "Operations → Evidence Repository search; upload naming preview via API.",
        "sources": "modules/operations/engines/evidence_repository.py; modules/audit_intelligence/engines/evidence_repository.py; app/evidence_routes.py; tests/test_usecase_batch1_evidence_workflows.py",
        "design": "Upload-time naming + tag metadata → repository search facets.",
        "remarks": "UI search is split across repository and legacy search modules.",
        "agenting": "No",
        "agenting_note": "Rule-based naming and tagging only.",
    },
    {
        "name": "Evidence dashboard and hash integrity check",
        "phase": "Phase1",
        "controls": "Cryptographic integrity validation, monitoring",
        "frameworks": FW,
        "why": "Provides operators a consolidated health view and integrity assurance for stored evidence.",
        "status": "Partially implemented",
        "justification": "Consolidated Evidence Dashboard (/mvp/evidence-dashboard) with Overview/Collection/Health tabs; integrity from get_health_dashboard() and audit pack verify_manifest; legacy /mvp/evidence-health retained.",
        "screen": "Operations → Evidence Dashboard (Overview, Collection, Health tabs); drill to Evidence Repository and Scheduler.",
        "sources": "modules/governance/templates/mvp_evidence_dashboard.html; modules/shared/routes/routes_mvp.py; modules/shared/services/module_capabilities.py; modules/operations/engines/evidence_repository.py; modules/governance/engines/evidence_health_engine.py; tests/test_evidence_navigation_refactor.py",
        "design": "Module capability aggregation → tabbed workspace → KPIs from analytics, scheduler, health and repository stats.",
        "remarks": "Health tab uses seeded health records (DEMO marker). Integrity KPI uses persisted repository rows when present.",
        "agenting": "No",
        "agenting_note": "Dashboard aggregation only; no autonomous remediation agent.",
    },
    {
        "name": "Common Control Library",
        "phase": "Phase1",
        "controls": "Cross-framework control mapping, audit documentation support",
        "frameworks": FW,
        "why": "Single reusable control catalog shared across PCI, DPSC, ISG, ISO and internal frameworks.",
        "status": "Implemented (MVP/demo)",
        "justification": "ECS_Query_Driven_Control_Library_Consolidated.xlsx (208 controls) plus supplementary_query_catalog.py; Phase-1 CommonControls/ folders (10 controls) with deterministic mock evidence collected by extended scheduler via common_controls_collector.py.",
        "screen": "Operations → Predefined Queries (catalog view); Administration → Scheduler (CommonControls collection); framework pages.",
        "sources": "ECS_Query_Driven_Control_Library_Consolidated.xlsx; modules/operations/engines/predefined_queries_engine.py; modules/operations/engines/common_controls_catalog.py; modules/operations/engines/common_controls_collector.py; CommonControls/; docs/use-cases/PREDEFINED_QUERY_INVENTORY.md; tests/test_common_controls_scheduler.py",
        "design": "Excel + code-defined supplementary entries → normalized control records → framework/technology indexes.",
        "remarks": "Assumption: Excel is authoritative; supplementary catalog is additive only. Not a separate nav item — surfaced via Predefined Queries and framework modules.",
        "agenting": "No",
        "agenting_note": "Static catalog data — no agent runtime.",
    },
    {
        "name": "Deterministic Predefined Queries",
        "phase": "Phase1",
        "controls": "Secure evidence retrieval, control completeness validation",
        "frameworks": FW,
        "why": "Runs technology-specific SQL/shell queries from the common library to collect deterministic evidence.",
        "status": "Implemented (MVP/demo)",
        "justification": "predefined_queries_engine loads 208 controls, assesses execution capability, exposes /mvp/predefined-queries UI and run endpoint; connector executors run allow-listed queries; tests in tests/test_predefined_query*.py and scripts/run_predefined_query_tests.py.",
        "screen": "Operations → Predefined Queries → filter catalog → Run → execution history.",
        "sources": "modules/operations/engines/predefined_queries_engine.py; modules/operations/templates/mvp_predefined_queries.html; modules/shared/routes/routes_mvp.py; modules/operations/engines/query_connectors.py; tests/test_predefined_query_execution.py",
        "design": "Control selection → connector dispatch → evidence registration → audit trail.",
        "remarks": "Live execution depends on configured DB/connector targets; demo mode may show capability assessment without live query.",
        "agenting": "No",
        "agenting_note": "Deterministic query execution — explicitly not LLM-generated SQL.",
    },
    {
        "name": "ECS Admin (users, roles, applications)",
        "phase": "Phase1",
        "controls": "RBAC, identity access management",
        "frameworks": FW,
        "why": "Administrative governance of personas, roles and onboarded applications.",
        "status": "Partially implemented",
        "justification": "admin_service with in-memory user registry (no secrets), /api/admin/* routes, /mvp/admin/users-roles UI; canonical roles from app.auth.roles.",
        "screen": "Administration paths → Users/Roles console (reachable by URL; not primary Phase-1 nav item).",
        "sources": "modules/shared/services/admin_service.py; modules/shared/routes/routes_mvp.py; modules/governance/templates/admin_users_roles.html; tests/test_usecase_batch1_evidence_workflows.py",
        "design": "RBAC-guarded admin API → in-memory user registry + read-only role/app catalogs.",
        "remarks": "User registry is demo/in-memory — not enterprise IdP provisioning.",
        "agenting": "No",
        "agenting_note": "CRUD admin surface only.",
    },
    {
        "name": "Evidence completeness detection",
        "phase": "Phase2",
        "controls": "Control completeness validation",
        "frameworks": FW,
        "why": "Identifies controls lacking evidence before audit cycles.",
        "status": "Partially implemented",
        "justification": "governance_completeness_engine and missing_evidence_engine power /mvp/completeness with gap export via comparison module.",
        "screen": "Governance → Completeness (Phase-2 nav / direct URL).",
        "sources": "modules/governance/engines/governance_completeness_engine.py; modules/governance/engines/missing_evidence_engine.py; modules/governance/templates/mvp_completeness.html; modules/shared/routes/routes_mvp.py",
        "design": "Framework/application control coverage → gap rows → export.",
        "remarks": "Uses catalog + seeded gap analytics; not all gaps are live CMDB-backed.",
        "agenting": "No",
        "agenting_note": "Analytics engine; no autonomous gap closure agent.",
    },
    {
        "name": "Evidence similarity and reuse",
        "phase": "Phase2",
        "controls": "Cross-framework control mapping",
        "frameworks": FW,
        "why": "Maps evidence once to satisfy multiple framework controls.",
        "status": "Partially implemented",
        "justification": "Reuse graph UI (/mvp/reuse, /mvp/evidence-story), evidence_reuse_service, pgvector RAG retrieval, rule-based reuse scoring.",
        "screen": "Operations → Evidence Reuse; reuse story and API /api/evidence-reuse/*.",
        "sources": "modules/governance/templates/mvp_reuse.html; modules/operations/templates/mvp_evidence_reuse_story.html; modules/audit_intelligence/services/evidence_reuse_service.py; ecs_platform/rag.py; app/evidence_intel/reuse.py",
        "design": "Repository records → reuse mapping / vector similarity → cross-framework suggestions.",
        "remarks": "Vector index population requires explicit reindex; not automatic on every upload.",
        "agenting": "No",
        "agenting_note": "Similarity is retrieval/scoring — not an agent workflow.",
    },
    {
        "name": "AI-generated evidence summaries",
        "phase": "Phase2",
        "controls": "Audit documentation support",
        "frameworks": FW,
        "why": "Summarises large evidence sets for auditors and control owners.",
        "status": "Partially implemented",
        "justification": "ai_ops_summary_engine, llm_engine provider abstraction, rag.answer with offline fallback modes; UI at /mvp/ai-ops-assistant/summary/{mode}.",
        "screen": "AI Ops Assistant → summary modes; copilot panel on workspace pages.",
        "sources": "modules/operations/engines/ai_ops_summary_engine.py; ecs_platform/llm_engine/provider.py; ecs_platform/rag.py; modules/operations/templates/mvp_ai_ops_assistant.html",
        "design": "Retrieve evidence context → LLM generate → cite/refuse per RAG policy.",
        "remarks": "Assumption: local Ollama default; cloud providers optional via config.",
        "agenting": "Yes",
        "agenting_note": "LLM orchestration with retrieval grounding — single-turn assistant pattern (not multi-agent planner).",
    },
    {
        "name": "Evidence quality scoring",
        "phase": "Phase2",
        "controls": "Evidence validation control",
        "frameworks": FW,
        "why": "Scores evidence artefacts for audit readiness and rejection risk.",
        "status": "Implemented (core), partial integration",
        "justification": "evidence_validation.compute validation with evidence_quality 0..1; surfaced via evidence_health and evidence_approval analytics; tests/test_evidence_validation.py.",
        "screen": "Evidence Dashboard Health tab; Evidence Approval Analytics; evidence-health drilldowns.",
        "sources": "modules/audit_intelligence/engines/evidence_validation.py; modules/governance/engines/evidence_health_engine.py; modules/governance/engines/evidence_approval_engine.py; tests/test_evidence_validation.py",
        "design": "Validation rules → quality score → health/approval KPIs.",
        "remarks": "Health view mixes scored records with seeded demo rows.",
        "agenting": "No",
        "agenting_note": "Rule-based scoring engine.",
    },
    {
        "name": "Natural language audit queries",
        "phase": "Phase2",
        "controls": "Secure evidence retrieval, compliance monitoring and governance reporting",
        "frameworks": FW,
        "why": "Lets auditors ask compliance questions in natural language with citations.",
        "status": "Implemented (with readiness dependency)",
        "justification": "rag.answer, chatbot_engine/enhanced, /chat and /api/platform/assistant; require_citations and refuse_without_evidence policies in RAG config.",
        "screen": "Global copilot; /mvp/ai-ops-assistant; LLM Prompt Workbench for prompt testing.",
        "sources": "ecs_platform/rag.py; modules/shared/services/chatbot_engine.py; modules/shared/services/chatbot_enhanced.py; modules/shared/templates/partials/chatbot_global.html",
        "design": "User query → retrieval → LLM answer with citations → fallback when no evidence.",
        "remarks": "Embedding index must be populated for best results; offline fallback modes exist.",
        "agenting": "Yes",
        "agenting_note": "RAG assistant pattern with tool-less retrieval grounding.",
    },
    {
        "name": "Leadership compliance dashboards",
        "phase": "Phase2",
        "controls": "Compliance monitoring and governance reporting",
        "frameworks": FW,
        "why": "Executive visibility into readiness, risk and open observations.",
        "status": "Implemented",
        "justification": "dashboard_service.executive_readiness, /mvp/audit/executive-readiness, /dashboard/cio, GET /api/audit/dashboard — verified in frontend_use_case_execution_verification.md.",
        "screen": "Overview dashboards; Audit Readiness executive view.",
        "sources": "modules/audit_intelligence/services/dashboard_service.py; modules/audit_intelligence/templates/audit/executive_readiness.html; modules/audit_intelligence/routes/routes_audit_ui.py; modules/executive_overview/templates/dashboard_cio.html",
        "design": "Aggregated readiness metrics → executive template → API drilldowns.",
        "remarks": "Some KPI tiles use demo/seeded enterprise metrics.",
        "agenting": "No",
        "agenting_note": "Read-only analytics dashboards.",
    },
    {
        "name": "Multi-application onboarding",
        "phase": "Phase3",
        "controls": "Application security baseline enforcement",
        "frameworks": FW,
        "why": "Onboards new banking applications into compliance scope with framework assignment.",
        "status": "Partially implemented",
        "justification": "onboarding_engine.simulate_onboarding, /mvp/onboarding UI with intake form and /api/onboarding/simulate; asset discovery hooks.",
        "screen": "Administration → Application Onboarding → intake → simulate/onboarder run.",
        "sources": "modules/operations/engines/onboarding_engine.py; modules/operations/templates/mvp_onboarding.html; modules/shared/routes/routes_mvp.py; tests/test_ui_nav_and_summary_cleanup.py",
        "design": "Intake form → simulate pipeline → framework assignment → CMDB-style records.",
        "remarks": "Simulation/demo pipeline — not full enterprise CMDB sync.",
        "agenting": "No",
        "agenting_note": "Wizard/simulation workflow.",
    },
    {
        "name": "Evidence lifecycle management",
        "phase": "Phase3",
        "controls": "Evidence retention, archival, version control",
        "frameworks": FW,
        "why": "Tracks evidence and observations from draft through approval and remediation.",
        "status": "Partially implemented",
        "justification": "Observation workflow transitions, ecs_platform governance review states, /mvp/lifecycle and platform evidence-lifecycle routes.",
        "screen": "Governance → Lifecycle views; observation transition API.",
        "sources": "modules/governance/engines/governance_lifecycle_engine.py; modules/governance/templates/mvp_lifecycle.html; modules/audit_intelligence/routes/routes_audit_intelligence.py",
        "design": "State machine (Draft→Submitted→Approved→…) → timeline views.",
        "remarks": "Not all lifecycle states wired to external ticketing.",
        "agenting": "No",
        "agenting_note": "Workflow state transitions — manual/API driven.",
    },
    {
        "name": "Cross-application compliance comparison",
        "phase": "Phase3",
        "controls": "Compliance monitoring and governance reporting",
        "frameworks": FW,
        "why": "Compares maturity and gaps across applications for prioritisation.",
        "status": "Implemented",
        "justification": "comparison_engine.build_comparison_dashboard, /mvp/comparison UI, gap export POST /mvp/comparison/export-gaps.",
        "screen": "Governance → Comparison (direct URL / Phase-2 nav when enabled).",
        "sources": "modules/governance/engines/comparison_engine.py; modules/governance/templates/mvp_comparison.html; modules/governance/engines/gap_export_engine.py",
        "design": "Application pairs → variance matrix → heatmap cards → export.",
        "remarks": "Demo readiness matrices seeded for stable CIO demos.",
        "agenting": "No",
        "agenting_note": "Analytics comparison only.",
    },
    {
        "name": "Automated control validation",
        "phase": "Phase3",
        "controls": "Evidence validation control, application security baseline enforcement",
        "frameworks": FW,
        "why": "Validates collected evidence against control expectations after ingestion.",
        "status": "Implemented (core), partial automation",
        "justification": "evidence_validation.validate_record, evidence_service.validate_run, /mvp/audit/validation UI; packs include verdict metadata.",
        "screen": "Audit validation results; pack build shows verdict per item.",
        "sources": "modules/audit_intelligence/engines/evidence_validation.py; modules/audit_intelligence/services/evidence_service.py; modules/audit_intelligence/templates/audit/validation_results.html; tests/test_evidence_validation.py",
        "design": "Run/evidence record → validation rules → verdict + quality score.",
        "remarks": "Automatic post-scheduler validation not fully chained in all paths.",
        "agenting": "No",
        "agenting_note": "Deterministic validation engine.",
    },
    {
        "name": "GRC platform integration",
        "phase": "Phase3",
        "controls": "Secure data transfer, encrypted storage",
        "frameworks": FW,
        "why": "Integrates ServiceNow, SharePoint and enterprise GRC sources for evidence collection.",
        "status": "Partially implemented",
        "justification": "Connector adapters (sharepoint_graph, servicenow_cmdb), integrations dashboard, connector test workbench with health/dry-run/parser-test for 11 connectors.",
        "screen": "Administration → Integrations; Connector Test Workbench; Evidence Explorer tab.",
        "sources": "modules/operations/templates/mvp_integrations.html; modules/audit_intelligence/services/connector_workbench.py; modules/integrations/sharepoint_graph.py; modules/integrations/servicenow_cmdb.py; tests/test_connector_execution_ingestion.py",
        "design": "Connector registry → health check → fetch/parse → evidence repository.",
        "remarks": "Live UAT requires credentials via env; mock parser-test path is deterministic.",
        "agenting": "No",
        "agenting_note": "Connector execution — not agent orchestration.",
    },
    {
        "name": "Enterprise compliance dashboards",
        "phase": "Phase3",
        "controls": "Compliance monitoring and governance reporting",
        "frameworks": FW,
        "why": "Organisation-wide compliance posture for CIO and compliance heads.",
        "status": "Partially implemented",
        "justification": "analytics_module.enterprise_dashboard, /mvp/enterprise template, executive_summary APIs; Phase-2 nav when show_phase2_nav enabled.",
        "screen": "Executive Overview → Enterprise (Phase-2 nav) or direct URL.",
        "sources": "modules/governance/engines/analytics_module.py; modules/executive_overview/templates/mvp_enterprise.html; modules/shared/routes/routes_mvp.py",
        "design": "Enterprise KPI aggregation → regional/BU breakdown templates.",
        "remarks": "Assumption: demo enterprise metrics — not live bank-wide feed.",
        "agenting": "No",
        "agenting_note": "Static/semi-seeded dashboards.",
    },
    {
        "name": "Automated regulatory reporting",
        "phase": "Phase4",
        "controls": "Audit documentation support, regulatory governance monitoring",
        "frameworks": FW,
        "why": "Generates audit-ready export packs for regulators and internal committees.",
        "status": "Partially implemented",
        "justification": "reporting_module + reports_analytics_engine; /mvp/reports with Evidence/Framework/Application/Audit/Observation/Evidence Packs tabs; download endpoints for PDF/Excel/CSV.",
        "screen": "Operations → Reports → tabbed export center → generate/download.",
        "sources": "modules/executive_overview/engines/reporting_module.py; modules/executive_overview/engines/reports_analytics_engine.py; modules/executive_overview/templates/mvp_reports.html; modules/shared/routes/routes_mvp.py",
        "design": "Report catalog → filtered tabs → export generator → download history.",
        "remarks": "Catalog is seeded; Evidence Packs tab reuses audit_repository_service.build_pack.",
        "agenting": "No",
        "agenting_note": "Template/report generation — no autonomous agent.",
    },
    {
        "name": "Cross-region audit analytics",
        "phase": "Phase4",
        "controls": "Compliance risk analytics",
        "frameworks": FW,
        "why": "Regional compliance analytics for distributed banking operations.",
        "status": "Partially implemented (demo analytics)",
        "justification": "Pan-India regional models in ecs_state and enterprise_mock_service; surfaced on /mvp/pan-india and enterprise views; largely seeded data.",
        "screen": "Executive Overview → Pan India (Phase-2 nav) — regional SLA and compliance tiles.",
        "sources": "modules/executive_overview/templates/mvp_pan_india.html; modules/executive_overview/engines/enterprise_mock_service.py; modules/shared/services/ecs_state.py",
        "design": "Regional seed dataset → aggregation cards → drill links.",
        "remarks": "Assumption: regional figures are demo/se seeded — not live core banking feeds.",
        "agenting": "No",
        "agenting_note": "Analytics only.",
    },
    {
        "name": "AI-assisted audit preparation",
        "phase": "Phase4",
        "controls": "Audit readiness automation",
        "frameworks": FW,
        "why": "Prepares audit packages and readiness checklists with AI assistance.",
        "status": "Partially implemented",
        "justification": "audit_preparation_checklist, /mvp/audit-prep, evidence_packs build/verify, LLM workbench for prompt testing; audit prep heatmaps.",
        "screen": "Governance → Audit Prep; Reports → Evidence Packs; Administration → LLM Prompt Workbench.",
        "sources": "modules/governance/engines/analytics_module.py; modules/governance/templates/mvp_audit_prep.html; modules/audit_intelligence/engines/evidence_packs.py; modules/audit_intelligence/templates/audit/llm_workbench.html",
        "design": "Readiness checklist → pack assembly → optional LLM summarisation via workbench/RAG.",
        "remarks": "Pack manifest verification is deterministic; LLM assists are optional/offline-capable.",
        "agenting": "Yes",
        "agenting_note": "Combines deterministic pack assembly with LLM prompt workflows for audit prep assistance.",
    },
    {
        "name": "Compliance trend analysis",
        "phase": "Phase4",
        "controls": "Compliance risk analytics",
        "frameworks": FW,
        "why": "Tracks closure velocity, rejections and aging for continuous improvement.",
        "status": "Implemented (analytics), demo-data dependent",
        "justification": "analytics_module.compliance_trends, trends_analytics_engine, /mvp/trends UI verified in use_case_implementation_matrix.md.",
        "screen": "Executive Overview → Trends (Phase-2 nav) or direct URL.",
        "sources": "modules/governance/engines/analytics_module.py; modules/governance/templates/mvp_trends.html; modules/executive_overview/engines/trends_analytics_engine.py",
        "design": "Historical trend series → closure/rejection/aging charts.",
        "remarks": "Trend series partly seeded for stable demo charts.",
        "agenting": "No",
        "agenting_note": "Analytics reporting.",
    },
    {
        "name": "National compliance dashboard",
        "phase": "Phase4",
        "controls": "Regulatory governance monitoring",
        "frameworks": FW,
        "why": "National-level compliance visibility for leadership committees.",
        "status": "Implemented (demo), partial production readiness",
        "justification": "mvp_pan_india.html, build_pan_india_posture, PAN_INDIA_REGIONS — frontend verified executable.",
        "screen": "Executive Overview → Pan India dashboard.",
        "sources": "modules/executive_overview/templates/mvp_pan_india.html; modules/executive_overview/engines/enterprise_mock_service.py; modules/shared/services/ecs_state.py",
        "design": "National region model → KPI strip → regional drill tables.",
        "remarks": "Assumption: national/regional KPIs are demo-seeded for budget/MVP presentations.",
        "agenting": "No",
        "agenting_note": "Executive dashboard — no agent layer.",
    },
]

CONTROL_MAPPINGS = [
    ("Evidence collection control, audit trail, access control", ["Manual evidence upload", "Bulk evidence upload"], "No", "Phase1", "Phase-1 reusable ingest controls shared across frameworks."),
    ("Encryption at rest, encryption in transit, automated evidence collection", ["Automated scheduled evidence pull"], "No", "Phase1", "Scheduler/connector collection control."),
    ("Evidence classification, control mapping", ["Metadata tagging and naming convention"], "No", "Phase1", "Shared metadata discipline for all frameworks."),
    ("Cryptographic integrity validation, monitoring", ["Evidence dashboard and hash integrity check"], "No", "Phase1", "SHA-256 integrity checks on repository artefacts."),
    ("Cross-framework control mapping, audit documentation support", ["Common Control Library", "Evidence similarity and reuse"], "Partial", "Phase1", "Library is shared; reuse maps across frameworks."),
    ("Secure evidence retrieval, control completeness validation", ["Deterministic Predefined Queries", "Natural language audit queries"], "Yes", "Phase1", "Predefined queries are deterministic; NL path uses RAG retrieval separately."),
    ("RBAC, identity access management", ["ECS Admin (users, roles, applications)"], "No", "Phase1", "Canonical roles and admin API."),
    ("Control completeness validation", ["Evidence completeness detection"], "No", "Phase2", "Gap detection against catalog."),
    ("Evidence validation control", ["Evidence quality scoring", "Automated control validation"], "Partial", "Phase2", "Validation engine shared; predefined query verdicts overlap."),
    ("Audit documentation support", ["AI-generated evidence summaries", "Automated regulatory reporting"], "No", "Phase2", "Summaries/reports support audit packs."),
    ("Compliance monitoring and governance reporting", ["Leadership compliance dashboards", "Cross-application compliance comparison", "Enterprise compliance dashboards"], "No", "Phase2", "Executive/monitoring controls."),
    ("Application security baseline enforcement", ["Multi-application onboarding", "Automated control validation"], "No", "Phase3", "Onboarding assigns baselines; validation confirms."),
    ("Evidence retention, archival, version control", ["Evidence lifecycle management"], "No", "Phase3", "Versioned audit repository + lifecycle states."),
    ("Secure data transfer, encrypted storage", ["GRC platform integration"], "No", "Phase3", "Connector transport to external GRC/evidence sources."),
    ("Audit readiness automation", ["AI-assisted audit preparation"], "Partial", "Phase4", "Pack builder + LLM-assisted prep flows."),
    ("Compliance risk analytics", ["Cross-region audit analytics", "Compliance trend analysis"], "No", "Phase4", "Regional and trend analytics."),
    ("Regulatory governance monitoring", ["National compliance dashboard", "Automated regulatory reporting"], "No", "Phase4", "National/regulatory reporting views."),
]

FRAMEWORK_COLS = ["PCI DSS", "DPSC", "ISG Baseline", "ISO 27001", "VAPT", "IS", "CSITE"]


def _fw_marks(control: str) -> list[str]:
    """Shared Phase-1 controls map to all major frameworks except VAPT where noted."""
    no_vapt = {
        "Evidence collection control, audit trail, access control",
        "Encryption at rest, encryption in transit, automated evidence collection",
        "Evidence classification, control mapping",
        "Cryptographic integrity validation, monitoring",
        "Cross-framework control mapping, audit documentation support",
        "Secure evidence retrieval, control completeness validation",
        "RBAC, identity access management",
        "Secure data transfer, encrypted storage",
        "Compliance monitoring and governance reporting",
        "Evidence retention, archival, version control",
        "Audit readiness automation",
        "Compliance risk analytics",
        "Regulatory governance monitoring",
        "Audit documentation support",
    }
    if control in no_vapt:
        return ["✓"] * 6 + ["–"]
    return ["✓"] * 7


def main() -> None:
    names = [u["name"] for u in USECASES]
    assert len(names) == len(set(names)), "duplicate use cases"

    for uc in USECASES:
        uc["sources"] = _exists(uc["sources"])

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1
    ws = wb.create_sheet("MD_Usecases")
    headers = [
        "Usecases", "Phase", "Controls Used", "Frameworks Using these Controls",
        "Why These Controls Are Required for the Use Case", "Implementation Status",
        "Implementation Justification", "Application Screen / Events",
        "Source Code Files / Routes", "Design Used", "Remarks / Pending Design",
        "Agenting Architecture (Yes/No)", "Agenting Justification",
    ]
    ws.append(headers)
    for h in ws[1]:
        h.font = Font(bold=True)
    for uc in sorted(USECASES, key=lambda u: (PHASE_ORDER[u["phase"]], u["name"])):
        ws.append([
            uc["name"], uc["phase"], uc["controls"], uc["frameworks"], uc["why"],
            uc["status"], uc["justification"], uc["screen"], uc["sources"],
            uc["design"], uc["remarks"], uc["agenting"], uc["agenting_note"],
        ])

    # Sheet 2
    ws2 = wb.create_sheet("Controls_Framework_Mapping")
    ws2.append(["Controls Used", *FRAMEWORK_COLS, "Use Cases", "Predefined Query (Yes/No)", "Phase", "Remarks"])
    for h in ws2[1]:
        h.font = Font(bold=True)
    for row in CONTROL_MAPPINGS:
        control, ucs, pq, phase, remarks = row
        ws2.append([control, *_fw_marks(control), ", ".join(ucs), pq, phase, remarks])

    # Summary counts
    counts = Counter(u["phase"] for u in USECASES)
    status_counts = Counter()
    for u in USECASES:
        s = u["status"]
        if s.startswith("Implemented") and "Partial" not in s and "partial" not in s:
            status_counts["Implemented"] += 1
        elif "Partial" in s or "partial" in s:
            status_counts["Partially implemented"] += 1
        else:
            status_counts["Other"] += 1

    # Sheet 3
    ws3 = wb.create_sheet("Implementation Summary")
    ws3.append(["ECS MD Use Cases — Source-Code Implementation Assessment"])
    ws3["A1"].font = Font(bold=True, size=12)
    ws3.append([])
    ws3.append(["Assessment basis", f"Direct inspection of ECS repository on {date.today().isoformat()} — code, docs/use-cases, and focused tests. Budget Approval / MVP stage; implementation only."])
    ws3.append(["Total use cases", len(USECASES)])
    ws3.append(["Phase1", counts["Phase1"]])
    ws3.append(["Phase2", counts["Phase2"]])
    ws3.append(["Phase3", counts["Phase3"]])
    ws3.append(["Phase4", counts["Phase4"]])
    ws3.append(["Implemented / core implemented", status_counts["Implemented"]])
    ws3.append(["Partially implemented / demo dependent", status_counts["Partially implemented"]])
    ws3.append(["Fully absent", 0])
    ws3.append([])
    ws3.append(["Interpretation", "Meaning", "MVP implication", "Assumption"])
    ws3.append(["Implemented", "End-to-end or core capability with reachable UI/API.", "Demonstrable in CIO/MVP demos with data-source caveats.", "Demo or seeded data may back some KPIs."])
    ws3.append(["Partially implemented", "Core engines/screens exist; live connectors or full automation incomplete.", "Show working surfaces; state what is dry-run/demo.", "Assumes ECS_CONNECTOR_EXECUTION_ENABLED=false by default."])
    ws3.append(["Agenting Yes", "LLM/RAG assistant participates in the use case.", "Disclose local/cloud LLM config dependency.", "Agenting = assistant/RAG pattern, not multi-agent swarm."])
    ws3.append(["Common Controls implemented", len(COMMON_CONTROLS), "Scheduler discovers CommonControls/ and persists deterministic mock evidence.", "Certificate Management demo evidence intentionally fails validation to produce an observation."])

    # Sheet 4 — Common Control Implementation
    ws_cc = wb.create_sheet("Common_Control_Implementation")
    cc_headers = [
        "Common Control", "Predefined Query", "Query IDs", "Alternate Collection",
        "Folder", "Scheduler", "Metadata", "Object Storage", "Validation",
        "Source Files", "Status",
    ]
    ws_cc.append(cc_headers)
    for h in ws_cc[1]:
        h.font = Font(bold=True)
    cc_engine_sources = _exists(
        "modules/operations/engines/common_controls_catalog.py; "
        "modules/operations/engines/common_controls_collector.py; "
        "modules/operations/engines/scheduler_module.py; "
        "tests/test_common_controls_scheduler.py"
    )
    for ctrl in COMMON_CONTROLS:
        folder = f"CommonControls/{ctrl.slug}/"
        folder_ok = (ROOT / folder).is_dir()
        sources = _exists(
            f"{folder}manifest.json; {folder}evidence.json; "
            "modules/operations/engines/common_controls_collector.py; "
            "modules/operations/engines/scheduler_module.py"
        )
        pq = "Yes" if ctrl.predefined_query_ids else "No"
        ws_cc.append([
            ctrl.name,
            pq,
            ", ".join(ctrl.predefined_query_ids),
            ctrl.alternate_collection,
            folder if folder_ok else "",
            "Extended run_scheduler_collection",
            "ai_repo.store_evidence + ops register_upload",
            "evidence_custody.resolve_custody (SNAPSHOT when enabled)",
            "Deterministic manifest rules (no AI)",
            sources,
            "Implemented (MVP/demo)",
        ])

    # Sheet 5
    ws4 = wb.create_sheet("Source Traceability")
    ws4.append(["Use Case", "Phase", "Status", "Primary UI/API", "Primary Source Files (verified)", "Implementation Design", "Verification / Pending Work"])
    for h in ws4[1]:
        h.font = Font(bold=True)
    for uc in sorted(USECASES, key=lambda u: (PHASE_ORDER[u["phase"]], u["name"])):
        ws4.append([uc["name"], uc["phase"], uc["status"], uc["screen"], uc["sources"], uc["design"], uc["remarks"]])
    # New source files only — Common Control Library implementation
    ws4.append([])
    ws4.append(["— Common Control Library (new sources) —", "Phase1", "Implemented (MVP/demo)",
                "Administration → Scheduler; CommonControls/",
                cc_engine_sources,
                "Scheduler discovers CommonControls/ → validate → persist → observation on fail",
                "Deterministic folder mock evidence; live predefined queries reused where catalog IDs exist"])
    for ctrl in COMMON_CONTROLS:
        folder_sources = _exists(
            f"CommonControls/{ctrl.slug}/manifest.json; CommonControls/{ctrl.slug}/evidence.json"
        )
        ws4.append([
            f"Common Control: {ctrl.name}",
            "Phase1",
            "Implemented (MVP/demo)",
            f"CommonControls/{ctrl.slug}/",
            folder_sources,
            "Folder manifest + evidence.json collected by common_controls_collector",
            f"Predefined queries: {', '.join(ctrl.predefined_query_ids) or 'none'}",
        ])

    # Sheet 6
    ws5 = wb.create_sheet("Assessment Notes")
    ws5.append(["Assessment Method and Important Caveats"])
    ws5["A1"].font = Font(bold=True, size=12)
    ws5.append([])
    notes = [
        ("Source reviewed", "Repository code (modules/*, app/*, ecs_platform/*), docs/use-cases/*.md, tests/test_evidence_navigation_refactor.py, tests/test_scheduler_run_wiring.py, tests/test_usecase_batch1_evidence_workflows.py."),
        ("Project stage", "Budget Approval / MVP — assesses implementation presence only; ignores HA, scalability, deployment maturity."),
        ("Meaning of Implemented", "Usable code path and reachable UI/API found. Does not imply live bank production data."),
        ("Meaning of Partial", "Core logic exists but may depend on demo data, dry-run defaults, or unconfigured connectors."),
        ("Common Control Library vs Predefined Queries", "Kept as separate use cases: library = catalog + CommonControls/ folder collection; predefined queries = deterministic execution layer."),
        ("Common Controls scheduler", "Extended existing scheduler only — discovers CommonControls/, validates deterministically, persists metadata/files, maps frameworks, generates observations on validation failure."),
        ("Evidence Dashboard", "Consolidated at /mvp/evidence-dashboard (Phase-1 nav); legacy /mvp/evidence-health and /mvp/evidence-approval routes retained."),
        ("Agenting Architecture", "Yes only where LLM/RAG assistant orchestration is part of the use case (summaries, NL queries, AI audit prep)."),
        ("Assumptions", "Seeded/demo data backs several dashboards; connector scheduler dry-run is default; CommonControls/ collection always runs on scheduler execute; 208 controls loaded from Excel; Certificate Management mock fails validation by design for MVP observation demo."),
        ("Recommended use", "Implementation reconciliation for budget approval — not production certification."),
    ]
    ws5.append(["Item", "Assessment note"])
    for item, note in notes:
        ws5.append([item, note])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print("Use cases:", len(USECASES), dict(counts))
    print("Status:", dict(status_counts))


if __name__ == "__main__":
    main()
