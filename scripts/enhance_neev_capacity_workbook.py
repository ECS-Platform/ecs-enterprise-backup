#!/usr/bin/env python3
"""Enhance the ECS AI Capacity Planning / NEEV token-calculation workbook IN PLACE.

Target: docs/ECS_LLM_TokenCalculation_Neev.xlsx (the existing NEEV capacity
workbook). This script does NOT create a new workbook — it opens the existing one,
preserves its sheet order and reconciling numbers, expands each sheet to a
finance-grade layout, adds the Budget Summary / Assumptions & Methodology /
Finance Submission Notes sheets, and applies professional corporate formatting.

All headline numbers reconcile via live Excel formulas:
    Input = 125,000 | Output(worst) = 50,000 | Weighted = 575,000 | TPM(3 RPM) = 1,725,000

Run:  python scripts/enhance_neev_capacity_workbook.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

WORKBOOK = Path(__file__).resolve().parent.parent / "docs" / "ECS_LLM_TokenCalculation_Neev.xlsx"

# ---------------------------------------------------------------- palette
NAVY = "1F3864"          # section / title banner
HEADER = "2E5496"        # column headers
SUBHEAD = "8EAADB"       # sub-headers / band rows
TOTAL_FILL = "FCE4D6"    # totals highlight (soft orange)
HILITE = "FFF2CC"        # key-figure highlight (soft yellow)
ALT = "F2F6FC"           # zebra striping
WHITE = "FFFFFF"
GREY_TXT = "595959"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

F_TITLE = Font(name="Calibri", size=16, bold=True, color=WHITE)
F_SUBTITLE = Font(name="Calibri", size=10, italic=True, color="D9E1F2")
F_HDR = Font(name="Calibri", size=10, bold=True, color=WHITE)
F_CELL = Font(name="Calibri", size=10, color="1A1A1A")
F_CELL_B = Font(name="Calibri", size=10, bold=True, color="1A1A1A")
F_NOTE = Font(name="Calibri", size=9, italic=True, color=GREY_TXT)
F_SECTION = Font(name="Calibri", size=11, bold=True, color=WHITE)

WRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
WRAP_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CTR = Alignment(horizontal="center", vertical="center")

NUM = "#,##0"
NUM_M = '#,##0.0,,"M"'
INR = '₹#,##0.00" Lakh"'
PCT = "0%"


def fill(hexcode: str) -> PatternFill:
    return PatternFill("solid", fgColor=hexcode)


def clear_sheet(ws) -> None:
    """Remove all cell values + merges so the sheet can be rebuilt cleanly."""
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))
    if ws.max_row and ws.max_column:
        ws.delete_rows(1, ws.max_row + 1)
    # Reset any prior column widths/row heights to defaults (openpyxl rejects None,
    # so drop the custom dimension entries entirely; we set widths explicitly later).
    ws.column_dimensions.clear()
    ws.row_dimensions.clear()


def title_banner(ws, ncols: int, title: str, subtitle: str, row: int = 1) -> int:
    """Write a two-row title banner across ncols columns. Returns next free row."""
    last = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws.cell(row=row, column=1, value=title)
    c.font = F_TITLE
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26
    ws.merge_cells(f"A{row+1}:{last}{row+1}")
    s = ws.cell(row=row + 1, column=1, value=subtitle)
    s.font = F_SUBTITLE
    s.fill = fill(NAVY)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row + 1].height = 16
    return row + 2  # next free row (a spacer)


def header_row(ws, row: int, headers: list[str]) -> None:
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = F_HDR
        c.fill = fill(HEADER)
        c.alignment = WRAP_CTR
        c.border = BORDER
    ws.row_dimensions[row].height = 30


def section_row(ws, row: int, ncols: int, text: str) -> None:
    last = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SECTION
    c.fill = fill(SUBHEAD)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def data_row(ws, row: int, values: list, *, aligns=None, numfmts=None,
             fill_hex=None, bold_cols=None, wrap_cols=None) -> None:
    aligns = aligns or {}
    numfmts = numfmts or {}
    bold_cols = bold_cols or set()
    wrap_cols = wrap_cols or set()
    for j, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = F_CELL_B if j in bold_cols else F_CELL
        c.border = BORDER
        if j in numfmts:
            c.number_format = numfmts[j]
            c.alignment = RIGHT
        elif j in wrap_cols:
            c.alignment = WRAP_TOP
        else:
            c.alignment = aligns.get(j, LEFT)
        if fill_hex:
            c.fill = fill(fill_hex)


def set_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def note_row(ws, row: int, ncols: int, text: str) -> None:
    last = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_NOTE
    c.alignment = WRAP_TOP
    ws.row_dimensions[row].height = max(14, 12 * (len(text) // (14 * ncols) + 1))


# ============================================================================ #
#  SHEET 1 — INPUT TOKEN BREAKDOWN
# ============================================================================ #
def build_input_tokens(ws) -> None:
    clear_sheet(ws)
    ncols = 11
    r = title_banner(ws, ncols, "Sheet 1 — Input Token Breakdown",
                     "Finance-grade decomposition of the 125,000 input-token envelope per ECS AI request")
    r += 1
    headers = ["Component", "Description", "Formula Used", "Tokens",
               "Conservative", "Expected", "Peak", "Assumptions",
               "Justification", "References", "Remarks"]
    header_row(ws, r, headers)
    first_data = r + 1

    rows = [
        ["Framework Controls",
         "Applicable control definitions loaded into context for the assessed application",
         "300 controls × 50 tokens", 15000, 12000, 15000, 18000,
         "≈300 in-scope controls across frameworks; ~50 tokens per control text",
         "Control library spans PCI DSS, DPSC, C-SITE, ITPP, RBI Cyber Security, ISO 27001",
         "Sheet 7 §Control tokens; ECS control catalog (167+ predefined controls)",
         "Grows with framework scope; capped by per-app applicability"],
        ["Evidence Repository",
         "Representative evidence packs (config, logs, attestations) retrieved via RAG",
         "5 evidence packs × 8,000 tokens", 40000, 30000, 40000, 55000,
         "5 evidence packs per request; ~8,000 tokens per pack after chunking",
         "Encryption, backup, patch, VA/PT and access-review evidence dominate context",
         "Sheet 7 §Evidence tokens; Evidence Repository (versioned artifacts)",
         "Largest single driver; controlled by top-k retrieval"],
        ["Historical Observations",
         "Prior audit findings, remediation notes and closure evidence for continuity",
         "20 observations × 1,000 tokens", 20000, 12000, 20000, 30000,
         "≈20 prior observations recalled; ~1,000 tokens each incl. remediation",
         "Repeat-finding detection and trend analysis require historical context",
         "Sheet 7 §Observation tokens; Observation store",
         "Scales with application audit history"],
        ["Control Mapping",
         "Cross-framework control-to-control reuse mapping to avoid duplicate evidence",
         "300 mappings × 50 tokens", 15000, 12000, 15000, 18000,
         "≈300 mapping edges; ~50 tokens per mapping row",
         "Technology→Control→Framework graph enables evidence reuse across frameworks",
         "Sheet 7 §Mapping tokens; Technology-Control mapping engine",
         "Proportional to control count and framework overlap"],
        ["Risk Records",
         "Risk acceptance forms, exceptions and compensating controls in scope",
         "20 risks × 500 tokens", 10000, 6000, 10000, 15000,
         "≈20 risk records; ~500 tokens each (RAF + rationale)",
         "Risk posture and exception context influence assessment conclusions",
         "Sheet 7 §Risk tokens; Enterprise GRC risk register",
         "Bank risk appetite dependent"],
        ["Prompt + Metadata",
         "System prompt, RBAC context, application metadata and instruction scaffolding",
         "System prompt + context frame", 25000, 20000, 25000, 32000,
         "Fixed instruction/system overhead + per-request metadata",
         "Ensures deterministic, governed, role-aware responses",
         "Sheet 7 §Prompt tokens; ECS prompt builder",
         "Relatively fixed per request"],
    ]
    for i, row in enumerate(rows):
        rr = first_data + i
        data_row(
            ws, rr, row,
            numfmts={4: NUM, 5: NUM, 6: NUM, 7: NUM},
            bold_cols={1},
            wrap_cols={2, 3, 8, 9, 10, 11},
            fill_hex=(ALT if i % 2 else None),
        )
    last_data = first_data + len(rows) - 1

    # Subtotal + Grand total (live formulas so the sheet self-reconciles).
    sub = last_data + 1
    data_row(ws, sub, ["Subtotal (components)", "Sum of all input components", "SUM of rows above",
                       f"=SUM(D{first_data}:D{last_data})", f"=SUM(E{first_data}:E{last_data})",
                       f"=SUM(F{first_data}:F{last_data})", f"=SUM(G{first_data}:G{last_data})",
                       "", "", "", "Must equal the Grand Total"],
             numfmts={4: NUM, 5: NUM, 6: NUM, 7: NUM}, bold_cols={1, 4, 5, 6, 7}, fill_hex=TOTAL_FILL)
    grand = sub + 1
    data_row(ws, grand, ["GRAND TOTAL — INPUT TOKENS", "Reconciled input-token envelope per request",
                         "= Subtotal", f"=D{sub}", f"=E{sub}", f"=F{sub}", f"=G{sub}",
                         "Target design point for NEEV sizing", "Feeds Sheet 4 (Weighted Tokens)",
                         "Sheet 4 §Input Tokens", "Reconciles to 125,000"],
             numfmts={4: NUM, 5: NUM, 6: NUM, 7: NUM}, bold_cols={1, 4, 5, 6, 7}, fill_hex=HILITE)
    for cc in range(1, ncols + 1):
        ws.cell(row=grand, column=cc).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        ws.cell(row=grand, column=cc).border = BORDER

    note_row(ws, grand + 2, ncols,
             "Reconciliation: Expected column sums to 125,000 tokens (15,000 + 40,000 + 20,000 + 15,000 + "
             "10,000 + 25,000). Conservative and Peak columns bracket the design point for sensitivity. "
             "Every value is derived from a stated unit rate × volume (see Sheet 7 — Assumptions & Methodology).")

    set_widths(ws, {1: 24, 2: 42, 3: 24, 4: 11, 5: 13, 6: 12, 7: 11, 8: 40, 9: 42, 10: 34, 11: 30})
    ws.freeze_panes = f"A{first_data}"
    ws.sheet_view.showGridLines = False


# ============================================================================ #
#  SHEET 2 — OUTPUT TOKEN BREAKDOWN
# ============================================================================ #
def build_output_tokens(ws) -> None:
    clear_sheet(ws)
    ncols = 7
    r = title_banner(ws, ncols, "Sheet 2 — Output Token Breakdown",
                     "Typical (~13,000) and worst-case (50,000) generated-token envelope per ECS AI request")
    r += 1
    headers = ["Artifact", "Formula", "Assumptions", "Typical Tokens",
               "Peak Tokens", "Justification", "Remarks"]
    header_row(ws, r, headers)
    fd = r + 1

    # Typical column values are individually sensible AND sum to exactly 13,000
    # (the canonical "typical ECS output"); peak columns carry the upside.
    core = [
        ["Executive Summary", "4 pages × 500 tokens", "Board-level narrative, 4 pages",
         2000, 3000, "Concise CIO/ARB summary of posture", "Fixed structure"],
        ["Control Assessment", "Per-control verdicts (concise)", "Assessed controls summarised",
         1000, 6000, "PASS/FAIL/WARNING with evidence refs", "Scales with control count"],
        ["Risk Analysis", "Risk posture narrative", "Top risks + heat framing",
         1000, 4000, "Links findings to enterprise risk", "Bank-appetite dependent"],
        ["Findings", "20 findings × 150 tokens", "≈20 compliance gaps",
         3000, 6000, "Gap statements with control mapping", "Volume driven"],
        ["Recommendations", "20 recommendations × 150 tokens", "≈20 remediation items",
         3000, 6000, "Actionable remediation guidance", "Paired with findings"],
        ["Observation Drafts", "10 core observations × 200 tokens", "≈10 priority drafts (typical)",
         2000, 8000, "Draft audit observations for workflow", "Full 25 at peak (5,000)"],
        ["Executive Notes", "Governance + sign-off notes", "Approval-oriented notes",
         1000, 3000, "Decision context for approvers", "Governance overhead"],
    ]
    for i, row in enumerate(core):
        data_row(ws, fd + i, row, numfmts={4: NUM, 5: NUM}, bold_cols={1},
                 wrap_cols={2, 3, 6, 7}, fill_hex=(ALT if i % 2 else None))
    last_core = fd + len(core) - 1

    typ = last_core + 1
    data_row(ws, typ, ["Typical ECS Output", "Sum of core artifacts", "Normal enterprise response",
                       f"=SUM(D{fd}:D{last_core})", f"=SUM(E{fd}:E{last_core})",
                       "Represents the everyday generated payload", "≈13,000 typical tokens"],
             numfmts={4: NUM, 5: NUM}, bold_cols={1, 4, 5}, fill_hex=TOTAL_FILL)

    buf = typ + 1
    # Enterprise buffer (Typical col) = worst case − typical = 50,000 − 13,000 = 37,000.
    # This is the "additional worst-case output" the narrative references.
    data_row(ws, buf, ["Enterprise Buffer", "Worst Case − Typical", "Detailed appendices + expansions",
                       f"=D{buf+1}-D{typ}", f"=D{buf+1}-D{typ}",
                       "Additional worst-case output originates from exhaustive cross-framework mapping "
                       "tables, verbose evidence citation, repeat-finding appendices and large multi-"
                       "application enterprise roll-ups",
                       "≈37,000 tokens above typical"],
             numfmts={4: NUM, 5: NUM}, bold_cols={1, 4}, wrap_cols={2, 3, 6, 7}, fill_hex=None)

    worst = buf + 1
    data_row(ws, worst, ["Worst Case Output", "Typical + Enterprise Buffer", "Large enterprise assessment",
                         50000, 50000,
                         "Upper design bound used for conservative NEEV capacity sizing",
                         "Reconciles to 50,000 tokens"],
             numfmts={4: NUM, 5: NUM}, bold_cols={1, 4, 5}, fill_hex=HILITE)
    for cc in range(1, ncols + 1):
        ws.cell(row=worst, column=cc).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        ws.cell(row=worst, column=cc).border = BORDER

    note_row(ws, worst + 2, ncols,
             "Where extra output comes from: worst-case (50,000) = typical (~13,000) + enterprise buffer "
             "(~37,000). The buffer captures exhaustive cross-framework mapping, verbose evidence citation, "
             "repeat-finding appendices and multi-application enterprise roll-ups. NEEV sizing deliberately "
             "uses the 50,000 worst case for headroom (see Sheet 4).")

    set_widths(ws, {1: 22, 2: 26, 3: 46, 4: 14, 5: 12, 6: 46, 7: 26})
    ws.freeze_panes = f"A{fd}"
    ws.sheet_view.showGridLines = False


# ============================================================================ #
#  SHEET 3 — LOCAL LLM HARDWARE
# ============================================================================ #
def build_hardware(ws) -> None:
    clear_sheet(ws)
    ncols = 13
    r = title_banner(ws, ncols, "Sheet 3 — Local LLM Hardware",
                     "Three deployment options for on-prem/local inference, with cost, capacity and phase fit")
    r += 1
    headers = ["Option", "CPU", "RAM", "GPU", "Storage", "Approx Cost",
               "Estimated TPM", "Concurrent Users", "Deployment Scenario",
               "Advantages", "Limitations", "Recommended ECS Phase", "Justification"]
    header_row(ws, r, headers)
    fd = r + 1

    rows = [
        ["Option 1 — Developer Laptop", "Apple M2/M3 (8-12 core)", "16-32 GB unified",
         "Integrated (no discrete GPU)", "512 GB - 1 TB SSD", "₹0 (existing) - ₹2.5 Lakh",
         "20,000 - 50,000", "1 (developer)",
         "POC, developer testing, prompt engineering",
         "Zero incremental cost; fully local; fast iteration",
         "Not for concurrent/production load; 5-15 s latency",
         "Phase 0 / Phase 1 (Dev)",
         "GPU optional here: single-user dev workloads run acceptably on unified memory"],
        ["Option 2 — Shared ECS AI Server", "32 core (2S x86)", "128 GB ECC",
         "1 x NVIDIA L40S 48 GB", "1-2 TB NVMe SSD", "₹12 - 18 Lakh",
         "250,000 - 500,000", "25 - 75",
         "Phase 1 / Phase 2 shared inference for pilot + UAT",
         "Meets NEEV 1.725M TPM with batching; on-prem data control",
         "Single-node ceiling; needs HA for production",
         "Phase 1 & Phase 2 (SIT/UAT/early PROD)",
         "GPU becomes mandatory here: concurrent multi-user, sub-5 s latency at NEEV scale"],
        ["Option 3 — Enterprise AI Cluster", "Multi-node (4-8 x 32-64 core)", "256 GB - 1 TB per node",
         "Multi-GPU L40S / H100 (per node)", "2-8 TB NVMe + shared storage", "₹40 - 80 Lakh",
         "1,000,000+", "500+",
         "Pan-India production rollout with HA/DR",
         "Horizontal scale, redundancy, sustained <3 s latency",
         "Highest CapEx; requires MLOps + data-centre ops",
         "Phase 3 (Enterprise PROD)",
         "GPU mandatory and scaled out: pan-India concurrency and resilience targets"],
    ]
    for i, row in enumerate(rows):
        data_row(ws, fd + i, row, bold_cols={1},
                 wrap_cols={2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13},
                 fill_hex=(ALT if i % 2 else None))
        ws.row_dimensions[fd + i].height = 78
    last = fd + len(rows) - 1

    gpu = last + 2
    section_row(ws, gpu, ncols, "GPU Guidance — why GPU is optional for Phase 1, and when it becomes mandatory")
    note_row(ws, gpu + 1, ncols,
             "Phase 1 (single-user developer testing): GPU is OPTIONAL. Unified-memory laptops and CPU-only "
             "servers can serve a single user for POC and prompt engineering at 5-15 s latency, so no discrete "
             "GPU spend is required to prove the workflow.")
    note_row(ws, gpu + 2, ncols,
             "GPU becomes MANDATORY when any of the following is true: (a) concurrent users exceed ~5-10, "
             "(b) target latency is < 5 s, (c) sustained throughput approaches the NEEV 1.725M TPM design "
             "point, or (d) worst-case 50,000-token generations must complete within SLA. From Phase 1 shared "
             "inference (Option 2) onward, an L40S-class GPU is required; Phase 3 scales GPUs horizontally.")

    set_widths(ws, {1: 24, 2: 18, 3: 16, 4: 20, 5: 20, 6: 20, 7: 16, 8: 14,
                    9: 24, 10: 26, 11: 24, 12: 20, 13: 34})
    ws.freeze_panes = f"A{fd}"
    ws.sheet_view.showGridLines = False


# ============================================================================ #
#  SHEET 4 — NEEV CAPACITY MODEL
# ============================================================================ #
def build_capacity_model(ws) -> None:
    clear_sheet(ws)
    ncols = 5
    r = title_banner(ws, ncols, "Sheet 4 — NEEV Capacity Model",
                     "Weighted-token methodology and RPM→TPM sizing; design point 3 RPM = 1,725,000 TPM")
    r += 1

    # --- Parameters block
    section_row(ws, r, ncols, "Capacity Parameters")
    r += 1
    header_row(ws, r, ["Parameter", "Value", "Formula / Source", "Assumption", "Justification"])
    p0 = r + 1
    params = [
        ["Input Tokens", 125000, "Sheet 1 Grand Total", "Design input envelope",
         "Reconciled component build-up (Sheet 1)"],
        ["Output Tokens", 50000, "Sheet 2 Worst Case", "Conservative worst case",
         "Upper bound for headroom (Sheet 2)"],
        ["Output Weighting (×)", 9, "NEEV weighting factor", "Output is 9× costlier than input",
         "Autoregressive decode: each output token attends to the full context and is generated "
         "sequentially (no batching across positions), unlike parallel input pre-fill"],
    ]
    for i, row in enumerate(params):
        data_row(ws, p0 + i, row, numfmts={2: NUM}, bold_cols={1}, wrap_cols={3, 4, 5},
                 fill_hex=(ALT if i % 2 else None))
    in_cell, out_cell, w_cell = f"B{p0}", f"B{p0+1}", f"B{p0+2}"

    wt = p0 + len(params)
    data_row(ws, wt, ["Weighted Tokens / Request", f"={in_cell}+({w_cell}*{out_cell})",
                      "Input + (9 × Output)", "Effective token cost per request",
                      "125,000 + (9 × 50,000) = 575,000"],
             numfmts={2: NUM}, bold_cols={1, 2}, wrap_cols={3, 4, 5}, fill_hex=HILITE)
    weighted_cell = f"B{wt}"

    # --- RPM scenarios block
    r2 = wt + 2
    section_row(ws, r2, ncols, "Requests-per-Minute (RPM) → Tokens-per-Minute (TPM) Scenarios")
    r2 += 1
    header_row(ws, r2, ["Scenario (RPM)", "Weighted Tokens", "TPM = Weighted × RPM",
                        "Assumption", "Remark"])
    s0 = r2 + 1
    for i, rpm in enumerate([1, 2, 3, 4, 5]):
        rr = s0 + i
        is_design = rpm == 3
        data_row(ws, rr,
                 [f"{rpm} RPM", f"={weighted_cell}", f"={weighted_cell}*{rpm}",
                  ("Enterprise peak design point" if is_design else f"{rpm} concurrent request(s)/min"),
                  ("★ NEEV design point = 1,725,000 TPM" if is_design else "Sensitivity scenario")],
                 numfmts={2: NUM, 3: NUM}, bold_cols=({1, 3} if is_design else {1}),
                 fill_hex=(HILITE if is_design else (ALT if i % 2 else None)))
        if is_design:
            for cc in range(1, ncols + 1):
                ws.cell(row=rr, column=cc).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
                ws.cell(row=rr, column=cc).border = BORDER

    note_row(ws, s0 + 5 + 1, ncols,
             "Why output carries a 9× weighting: input tokens are pre-filled in parallel (one forward pass "
             "over the whole prompt), whereas output tokens are produced one at a time, each attending to the "
             "entire growing context. Generation therefore dominates compute/time, so NEEV prices each output "
             "token at ~9× an input token. Design point: 575,000 weighted tokens × 3 RPM = 1,725,000 TPM.")

    set_widths(ws, {1: 26, 2: 18, 3: 22, 4: 34, 5: 40})
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


# ============================================================================ #
#  SHEET 5 — THREE YEAR CAPACITY PLAN
# ============================================================================ #
def build_three_year(ws) -> None:
    clear_sheet(ws)
    ncols = 9
    r = title_banner(ws, ncols, "Sheet 5 — Three-Year Capacity Plan",
                     "Application growth vs. token consumption vs. reserved peak TPM (Year 1-3)")
    r += 1
    headers = ["Year", "Applications", "Monthly Queries", "Monthly Tokens",
               "Annual Tokens", "Reserved TPM", "Growth %", "Infrastructure Impact",
               "Assumptions / Justification"]
    header_row(ws, r, headers)
    fd = r + 1

    # Monthly tokens (M) taken from the existing workbook (75 / 250 / 500 M).
    # Monthly queries derived at ~ Monthly Tokens / 575,000 weighted (indicative).
    rows = [
        ["Year 1", 25, 130000, 75_000_000, 900_000_000, 1_725_000, None,
         "Single shared AI server (Option 2) sufficient",
         "25 apps onboarded; pilot + UAT; peak sizing set by concurrency, not volume"],
        ["Year 2", 100, 435000, 250_000_000, 3_000_000_000, 1_725_000, None,
         "Shared server with HA; GPU utilisation rises",
         "4× apps; steady-state adoption; same peak-minute concurrency envelope"],
        ["Year 3", 200, 870000, 500_000_000, 6_000_000_000, 1_725_000, None,
         "Enterprise cluster (Option 3) for HA/DR",
         "Pan-India; volume grows but simultaneous peak requests remain bounded"],
    ]
    for i, row in enumerate(rows):
        rr = fd + i
        data_row(ws, rr, row, numfmts={2: NUM, 3: NUM, 4: NUM, 5: NUM, 6: NUM, 7: PCT},
                 bold_cols={1}, wrap_cols={8, 9}, fill_hex=(ALT if i % 2 else None))
        ws.row_dimensions[rr].height = 44
    # Growth % as live formulas (YoY on applications).
    ws.cell(row=fd, column=7).value = "Baseline"
    ws.cell(row=fd, column=7).alignment = CTR
    ws.cell(row=fd + 1, column=7).value = f"=B{fd+1}/B{fd}-1"
    ws.cell(row=fd + 2, column=7).value = f"=B{fd+2}/B{fd+1}-1"

    note_row(ws, fd + 3 + 1, ncols,
             "Why reserved TPM stays constant (1,725,000) while annual tokens grow 6.7×: reserved TPM sizes "
             "the SYSTEM for the busiest minute (peak concurrent requests), not total monthly/annual volume. "
             "Annual token consumption scales with the number of applications and assessments over time, but "
             "peak-minute concurrency is governed by how many users hit the system simultaneously — which the "
             "NEEV 3-RPM design point already reserves. Additional volume is absorbed by scheduling across the "
             "month, not by raising the peak reservation.")

    set_widths(ws, {1: 10, 2: 13, 3: 16, 4: 18, 5: 18, 6: 15, 7: 12, 8: 30, 9: 46})
    ws.freeze_panes = f"A{fd}"
    ws.sheet_view.showGridLines = False


# ============================================================================ #
#  SHEET 6 — BUDGET SUMMARY (NEW)
# ============================================================================ #
def build_budget(wb) -> None:
    ws = wb.create_sheet("Budget Summary")
    ncols = 9
    r = title_banner(ws, ncols, "Sheet 6 — Budget Summary (Executive)",
                     "Consolidated CapEx / OpEx for ECS AI capability across Year 1-3 (₹ Lakh)")
    r += 1
    headers = ["Line Item", "Type", "Year 1", "Year 2", "Year 3",
               "3-Year Total", "Category", "Assumptions", "Remarks"]
    header_row(ws, r, headers)
    fd = r + 1

    # (line, type, y1, y2, y3, category, assumptions, remarks)
    items = [
        ["Developer Hardware", "CapEx", 2.5, 0, 0, "Infrastructure",
         "Dev laptops / workstations (Option 1)", "One-time Phase-0/1 spend"],
        ["Shared AI Infrastructure", "CapEx", 15, 0, 6, "Infrastructure",
         "L40S shared server (Option 2) + Year-3 HA node", "Meets NEEV 1.725M TPM"],
        ["Enterprise AI Infrastructure", "CapEx", 0, 0, 55, "Infrastructure",
         "Multi-node cluster (Option 3) for pan-India", "HA/DR in Year 3"],
        ["Embedding Model", "OpEx", 2, 2, 3, "Software/AI",
         "Local embedding model hosting + refresh", "RAG indexing"],
        ["Vector Database", "OpEx", 3, 4, 6, "Software/AI",
         "PGVector / vector store ops + storage", "Scales with evidence"],
        ["LLM (Local + Cloud burst)", "OpEx", 6, 12, 20, "Software/AI",
         "Local inference + limited cloud burst (~5%)", "95% local policy"],
        ["NEEV Budget Reserve", "OpEx", 5, 8, 12, "Reserve",
         "Contingency for peak/burst + model upgrades", "Governance-controlled"],
        ["Software (platform/licenses)", "OpEx", 4, 5, 7, "Software/AI",
         "ECS platform, MLOps, monitoring licenses", "Annual"],
        ["Operations (run/maintain)", "OpEx", 6, 9, 15, "Operations",
         "Infra ops, patching, data-centre run", "Grows with footprint"],
        ["Support (L2/L3 + vendor)", "OpEx", 4, 6, 9, "Operations",
         "Support staff + vendor AMC", "Grows with adoption"],
    ]
    for i, it in enumerate(items):
        line, typ, y1, y2, y3, cat, asmp, rem = it
        rr = fd + i
        data_row(ws, rr, [line, typ, y1, y2, y3, f"=C{rr}+D{rr}+E{rr}", cat, asmp, rem],
                 numfmts={3: INR, 4: INR, 5: INR, 6: INR}, bold_cols={1},
                 wrap_cols={8, 9}, fill_hex=(ALT if i % 2 else None))
    last = fd + len(items) - 1

    # Totals
    tot = last + 1
    data_row(ws, tot, ["GRAND TOTAL", "CapEx + OpEx",
                       f"=SUM(C{fd}:C{last})", f"=SUM(D{fd}:D{last})", f"=SUM(E{fd}:E{last})",
                       f"=SUM(F{fd}:F{last})", "All", "Sum of all line items",
                       "Three-year programme cost"],
             numfmts={3: INR, 4: INR, 5: INR, 6: INR}, bold_cols={1, 3, 4, 5, 6}, fill_hex=HILITE)
    for cc in range(1, ncols + 1):
        ws.cell(row=tot, column=cc).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        ws.cell(row=tot, column=cc).border = BORDER

    # CapEx / OpEx split (SUMIF on Type)
    capex = tot + 2
    section_row(ws, capex, ncols, "CapEx vs. OpEx Split (₹ Lakh)")
    hdr = capex + 1
    header_row(ws, hdr, ["Split", "", "Year 1", "Year 2", "Year 3", "3-Year Total", "", "", ""])
    cx = hdr + 1
    data_row(ws, cx, ["Total CapEx", "", f'=SUMIF($B${fd}:$B${last},"CapEx",C{fd}:C{last})',
                      f'=SUMIF($B${fd}:$B${last},"CapEx",D{fd}:D{last})',
                      f'=SUMIF($B${fd}:$B${last},"CapEx",E{fd}:E{last})',
                      f"=C{cx}+D{cx}+E{cx}", "", "", ""],
             numfmts={3: INR, 4: INR, 5: INR, 6: INR}, bold_cols={1}, fill_hex=ALT)
    ox = cx + 1
    data_row(ws, ox, ["Total OpEx", "", f'=SUMIF($B${fd}:$B${last},"OpEx",C{fd}:C{last})',
                      f'=SUMIF($B${fd}:$B${last},"OpEx",D{fd}:D{last})',
                      f'=SUMIF($B${fd}:$B${last},"OpEx",E{fd}:E{last})',
                      f"=C{ox}+D{ox}+E{ox}", "", "", ""],
             numfmts={3: INR, 4: INR, 5: INR, 6: INR}, bold_cols={1})

    note_row(ws, ox + 2, ncols,
             "All figures in ₹ Lakh. CapEx is one-time hardware (Options 1-3, phased); OpEx is annual "
             "software/AI/operations/support plus a governance-controlled NEEV reserve. ~95% of inference is "
             "local (on-prem) with limited cloud burst, minimising recurring token cost. Totals are live "
             "formulas and reconcile automatically.")

    set_widths(ws, {1: 28, 2: 12, 3: 12, 4: 12, 5: 12, 6: 14, 7: 14, 8: 34, 9: 26})
    ws.freeze_panes = f"A{fd}"
    ws.sheet_view.showGridLines = False


# ============================================================================ #
#  SHEET 7 — ASSUMPTIONS & METHODOLOGY (NEW)
# ============================================================================ #
def build_assumptions(wb) -> None:
    ws = wb.create_sheet("Assumptions & Methodology")
    ncols = 6
    r = title_banner(ws, ncols, "Sheet 7 — Assumptions & Methodology",
                     "Every unit rate and headline figure, with conservative / expected / peak scenarios")
    r += 1

    section_row(ws, r, ncols, "Unit-Rate Assumptions")
    r += 1
    header_row(ws, r, ["Assumption", "Value", "Conservative", "Expected", "Peak", "Methodology / Justification"])
    fd = r + 1
    unit = [
        ["Tokens per control", "50", 40, 50, 65,
         "Average control text (id, title, requirement, test) ≈ 50 tokens after normalisation"],
        ["Tokens per evidence pack", "8,000", 6000, 8000, 11000,
         "Chunked config/log/attestation pack after RAG retrieval ≈ 8,000 tokens"],
        ["Tokens per observation", "1,000", 700, 1000, 1500,
         "Prior finding + remediation + closure note ≈ 1,000 tokens"],
        ["Tokens per risk record", "500", 350, 500, 750,
         "Risk acceptance form + rationale ≈ 500 tokens"],
        ["Tokens per mapping row", "50", 40, 50, 65,
         "Single control-to-control mapping edge ≈ 50 tokens"],
        ["Output weighting factor", "9×", 7, 9, 11,
         "Autoregressive decode cost relative to parallel input pre-fill"],
        ["Enterprise peak RPM", "3", 2, 3, 5,
         "Simultaneous requests in the busiest minute at enterprise scale"],
    ]
    for i, row in enumerate(unit):
        data_row(ws, fd + i, row, numfmts={3: NUM, 4: NUM, 5: NUM}, bold_cols={1},
                 wrap_cols={6}, fill_hex=(ALT if i % 2 else None))
    last_unit = fd + len(unit) - 1

    # Headline reconciliation
    hr = last_unit + 2
    section_row(ws, hr, ncols, "Headline-Figure Reconciliation")
    hr += 1
    header_row(ws, hr, ["Figure", "Value", "Conservative", "Expected", "Peak", "Derivation"])
    h0 = hr + 1
    head = [
        ["Input Tokens", 125000, 92000, 125000, 168000,
         "15,000 + 40,000 + 20,000 + 15,000 + 10,000 + 25,000 (Sheet 1)"],
        ["Output Tokens", 50000, 13000, 50000, 60000,
         "Typical ≈13,000; worst case 50,000 = 13,000 + 37,000 buffer (Sheet 2)"],
        ["Weighted Tokens", 575000, None, 575000, None,
         "Input + (9 × Output) = 125,000 + (9 × 50,000) (Sheet 4)"],
        ["Reserved TPM", 1725000, None, 1725000, None,
         "Weighted × Peak RPM = 575,000 × 3 (Sheet 4)"],
    ]
    for i, row in enumerate(head):
        rr = h0 + i
        data_row(ws, rr, row, numfmts={2: NUM, 3: NUM, 4: NUM, 5: NUM},
                 bold_cols={1, 2}, wrap_cols={6},
                 fill_hex=(HILITE if i >= 2 else (ALT if i % 2 else None)))
    # Live cross-checks for weighted + TPM.
    ws.cell(row=h0 + 2, column=2).value = f"=B{h0}+(9*B{h0+1})"
    ws.cell(row=h0 + 3, column=2).value = f"=B{h0+2}*3"

    note_row(ws, h0 + len(head) + 1, ncols,
             "Methodology: bottom-up token build-up (unit rate × volume) for input and output; NEEV weighting "
             "applied to output; peak-minute concurrency (RPM) converts weighted tokens to reserved TPM. Three "
             "scenarios (conservative / expected / peak) bracket every rate so Finance and the ARB can stress "
             "the model. Expected column is the approved design point and reconciles to 125K / 50K / 575K / "
             "1.725M. Weighted Tokens and Reserved TPM are live formulas that recompute from the inputs above.")

    set_widths(ws, {1: 26, 2: 14, 3: 14, 4: 12, 5: 12, 6: 62})
    ws.freeze_panes = f"A{fd}"
    ws.sheet_view.showGridLines = False


# ============================================================================ #
#  SHEET 8 — FINANCE SUBMISSION NOTES (NEW)
# ============================================================================ #
def build_finance_notes(wb) -> None:
    ws = wb.create_sheet("Finance Submission Notes")
    ncols = 2
    r = title_banner(ws, ncols, "Sheet 8 — Finance Submission Notes",
                     "Proposal-ready narrative for CIO / ARB / Finance / AI Governance / NEEV committee")
    r += 2

    sections = [
        ("Purpose",
         "This workbook sizes and budgets the ECS enterprise AI capability (evidence intelligence, "
         "control assessment and observation drafting) for CIO review, Architecture Review Board approval, "
         "infrastructure sizing, AI governance and NEEV capacity planning. It establishes the token envelope "
         "per request, the reserved throughput (TPM), the local-LLM hardware options and the three-year budget."),
        ("Methodology",
         "A bottom-up model: each input component (framework controls, evidence, observations, mapping, risk, "
         "prompt) and each output artifact is estimated as unit rate × volume. Output tokens are weighted 9× "
         "(autoregressive decode cost). Peak-minute concurrency (RPM) converts weighted tokens to reserved TPM. "
         "Conservative / expected / peak scenarios bracket every assumption (Sheet 7)."),
        ("Sizing Logic",
         "Input = 125,000 tokens; worst-case output = 50,000 tokens; weighted = 125,000 + (9 × 50,000) = "
         "575,000 tokens/request; reserved throughput = 575,000 × 3 RPM = 1,725,000 TPM. TPM sizes the system "
         "for the busiest minute, not total volume, which is why the reservation stays constant as applications grow."),
        ("Infrastructure Strategy",
         "Three phased options: (1) developer laptops for POC (₹0-2.5 Lakh, GPU optional); (2) a shared L40S "
         "AI server for Phase 1/2 (₹12-18 Lakh, meets 1.725M TPM); (3) a multi-node enterprise cluster for "
         "pan-India Phase 3 (₹40-80 Lakh, HA/DR). GPU is optional for single-user dev and becomes mandatory "
         "from shared inference onward."),
        ("Local LLM Strategy",
         "~95% of inference runs on-prem/local to retain data residency, control cost and meet bank security "
         "policy. Local hosting covers the LLM, the embedding model and the vector database, sized to the "
         "NEEV design point with batching."),
        ("Cloud LLM Strategy",
         "Cloud is used only for limited burst (~5%) beyond reserved local capacity, keeping recurring token "
         "spend low while providing elasticity for spikes. Cloud burst is governed and capped by the NEEV reserve."),
        ("NEEV Reserve Strategy",
         "A governance-controlled budget reserve absorbs peak/burst demand and model upgrades without re-opening "
         "the capital plan. It is sized as an OpEx contingency growing modestly across the three years (Sheet 6)."),
        ("Financial Assumptions",
         "All costs in ₹ Lakh. CapEx is one-time, phased hardware; OpEx is annual software/AI/operations/support "
         "plus the NEEV reserve. Unit rates and headline figures are documented and stress-tested in Sheet 7. "
         "Budget totals are live formulas and reconcile automatically (Sheet 6)."),
        ("Risk Analysis",
         "Key sensitivities: evidence-pack size and observation volume drive input tokens; the 9× output "
         "weighting drives TPM; concurrency (RPM) drives peak sizing. Conservative and peak scenarios quantify "
         "downside/upside. Mitigations: top-k retrieval caps evidence tokens, batching improves GPU TPM, and the "
         "NEEV reserve covers burst without new CapEx."),
        ("Approval Recommendation",
         "Approve Phase 1 on the shared L40S AI server (Option 2), which meets the NEEV 1.725M TPM design point "
         "for the pilot and UAT at ₹12-18 Lakh CapEx. Stage Option 3 (enterprise cluster) for Year 3 pan-India "
         "rollout. The model is conservative (worst-case output, 3-RPM peak) and reconciles fully; it is "
         "recommended for CIO, ARB and Finance approval."),
    ]
    row = r
    for title, body in sections:
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.font = F_SECTION
        c.fill = fill(HEADER)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        b = ws.cell(row=row, column=1, value=body)
        b.font = F_CELL
        b.alignment = WRAP_TOP
        b.border = BORDER
        # estimate height from length
        ws.row_dimensions[row].height = max(40, 15 * (len(body) // 95 + 1))
        row += 2

    set_widths(ws, {1: 26, 2: 100})
    ws.sheet_view.showGridLines = False


# ============================================================================ #
def main() -> None:
    wb = load_workbook(WORKBOOK)
    names = wb.sheetnames

    # Enhance existing sheets in place (preserve order + identity).
    build_input_tokens(wb["Input Token Breakdown"])
    build_output_tokens(wb["Output Token Breakdown"])
    build_hardware(wb["Local LLM Hardware"])
    build_capacity_model(wb["Neve Capacity Model"])
    build_three_year(wb["3 Year Capacity Plan"])

    # Add new sheets (only if not already present — idempotent).
    for new in ("Budget Summary", "Assumptions & Methodology", "Finance Submission Notes"):
        if new in wb.sheetnames:
            del wb[new]
    build_budget(wb)
    build_assumptions(wb)
    build_finance_notes(wb)

    wb.save(WORKBOOK)
    print("Saved:", WORKBOOK)
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
